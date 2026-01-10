"""Tests for Excel parser adapter."""

import io

import pytest
from openpyxl import Workbook

from src.glp.assignment.adapters.excel_parser import OpenpyxlExcelParser


@pytest.fixture
def parser():
    return OpenpyxlExcelParser()


@pytest.fixture
def sample_excel_bytes():
    """Create a sample Excel file in memory."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Devices"

    # Headers
    ws["A1"] = "Serial Number"
    ws["B1"] = "MAC Address"

    # Data
    ws["A2"] = "SN12345"
    ws["B2"] = "00:1B:44:11:3A:B7"
    ws["A3"] = "SN67890"
    ws["B3"] = "AA:BB:CC:DD:EE:FF"
    ws["A4"] = "SN11111"
    # No MAC address for row 4

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


@pytest.fixture
def excel_with_different_headers():
    """Excel with alternative column names."""
    wb = Workbook()
    ws = wb.active

    ws["A1"] = "serial"  # Alternative name
    ws["B1"] = "mac"  # Alternative name

    ws["A2"] = "SN999"
    ws["B2"] = "11:22:33:44:55:66"

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


class TestOpenpyxlExcelParser:
    """Tests for OpenpyxlExcelParser."""

    def test_parse_valid_excel(self, parser, sample_excel_bytes):
        rows = parser.parse(sample_excel_bytes)

        assert len(rows) == 3

        assert rows[0].serial_number == "SN12345"
        assert rows[0].mac_address == "00:1B:44:11:3A:B7"
        assert rows[0].row_number == 2

        assert rows[1].serial_number == "SN67890"
        assert rows[1].mac_address == "AA:BB:CC:DD:EE:FF"

        assert rows[2].serial_number == "SN11111"
        assert rows[2].mac_address is None

    def test_parse_alternative_headers(self, parser, excel_with_different_headers):
        rows = parser.parse(excel_with_different_headers)

        assert len(rows) == 1
        assert rows[0].serial_number == "SN999"
        assert rows[0].mac_address == "11:22:33:44:55:66"

    def test_parse_empty_file(self, parser):
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "Serial Number"

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        rows = parser.parse(output.getvalue())
        assert len(rows) == 0

    def test_parse_missing_serial_column_raises(self, parser):
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "Wrong Column"
        ws["A2"] = "Value"

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        with pytest.raises(ValueError, match="Serial Number"):
            parser.parse(output.getvalue())

    def test_parse_invalid_file_raises(self, parser):
        with pytest.raises(ValueError, match="Failed to parse"):
            parser.parse(b"not an excel file")

    def test_validate_valid_rows(self, parser, sample_excel_bytes):
        rows = parser.parse(sample_excel_bytes)
        result = parser.validate(rows)

        assert result.is_valid is True
        assert len(result.errors) == 0

    def test_validate_duplicate_serials(self, parser):
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "Serial Number"
        ws["A2"] = "SN123"
        ws["A3"] = "SN123"  # Duplicate

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        rows = parser.parse(output.getvalue())
        result = parser.validate(rows)

        assert result.is_valid is False
        assert len(result.errors) == 1
        assert "Duplicate" in result.errors[0].message

    def test_validate_invalid_mac_format(self, parser):
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "Serial Number"
        ws["B1"] = "MAC Address"
        ws["A2"] = "SN123"
        ws["B2"] = "invalid-mac"

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        rows = parser.parse(output.getvalue())
        result = parser.validate(rows)

        assert result.is_valid is False
        assert len(result.errors) == 1
        assert "MAC" in result.errors[0].message

    def test_normalize_mac_formats(self, parser):
        """Test that various MAC formats are normalized."""
        # Test the static method directly
        assert parser._normalize_mac("aabbccddeeff") == "AA:BB:CC:DD:EE:FF"
        assert parser._normalize_mac("AA-BB-CC-DD-EE-FF") == "AA:BB:CC:DD:EE:FF"
        assert parser._normalize_mac("AABB.CCDD.EEFF") == "AA:BB:CC:DD:EE:FF"

    def test_large_file_warning(self, parser):
        """Test that large files generate a warning."""
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "Serial Number"

        # Add 1001 rows
        for i in range(2, 1003):
            ws[f"A{i}"] = f"SN{i:05d}"

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        rows = parser.parse(output.getvalue())
        result = parser.validate(rows)

        assert result.is_valid is True
        assert len(result.warnings) == 1
        assert "Large file" in result.warnings[0]
