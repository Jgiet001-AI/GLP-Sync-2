"""Excel and CSV parser adapter.

This adapter implements IExcelParser to parse Excel and CSV files
containing device serial numbers and MAC addresses.
"""

import csv
import io
import logging
import re
from typing import Optional

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from ..domain.entities import ExcelRow, ValidationError, ValidationResult
from ..domain.ports import IExcelParser

logger = logging.getLogger(__name__)

# MAC address regex pattern (accepts various formats)
MAC_PATTERN = re.compile(
    r"^([0-9A-Fa-f]{2}[:-]){5}[0-9A-Fa-f]{2}$|"  # XX:XX:XX:XX:XX:XX or XX-XX-XX-XX-XX-XX
    r"^([0-9A-Fa-f]{4}\.){2}[0-9A-Fa-f]{4}$|"  # XXXX.XXXX.XXXX
    r"^[0-9A-Fa-f]{12}$"  # XXXXXXXXXXXX
)


class OpenpyxlExcelParser(IExcelParser):
    """Excel parser implementation using openpyxl.

    Expected Excel format:
    | Serial Number | MAC Address       |
    |---------------|-------------------|
    | SN12345       | 00:1B:44:11:3A:B7 |
    | SN67890       |                   |

    - First row is treated as header
    - Serial Number column is required
    - MAC Address column is optional
    """

    # Column name variations we accept
    SERIAL_COLUMNS = ["serial number", "serial", "serialnumber", "serial_number", "sn"]
    MAC_COLUMNS = ["mac address", "mac", "macaddress", "mac_address"]

    def parse(self, file_content: bytes) -> list[ExcelRow]:
        """Parse an Excel or CSV file.

        Args:
            file_content: Raw bytes of the Excel or CSV file

        Returns:
            List of ExcelRow objects

        Raises:
            ValueError: If file format is invalid
        """
        # Try to detect if it's a CSV file
        if self._is_csv(file_content):
            return self._parse_csv(file_content)
        return self._parse_excel(file_content)

    def _is_csv(self, file_content: bytes) -> bool:
        """Detect if file content is CSV format.

        Args:
            file_content: Raw bytes of the file

        Returns:
            True if CSV, False otherwise
        """
        try:
            # Try to decode as text - CSV files are text-based
            text = file_content.decode("utf-8-sig")  # Handle BOM
            # Check if it looks like CSV (has commas/semicolons and newlines)
            if "\n" in text or "\r" in text:
                first_line = text.split("\n")[0].split("\r")[0]
                # CSV typically has delimiters
                if "," in first_line or ";" in first_line or "\t" in first_line:
                    return True
        except UnicodeDecodeError:
            # Not a text file, likely Excel
            pass
        return False

    def _parse_csv(self, file_content: bytes) -> list[ExcelRow]:
        """Parse a CSV file.

        Args:
            file_content: Raw bytes of the CSV file

        Returns:
            List of ExcelRow objects

        Raises:
            ValueError: If file format is invalid
        """
        try:
            # Decode content
            text = file_content.decode("utf-8-sig")  # Handle BOM

            # Detect delimiter
            sniffer = csv.Sniffer()
            try:
                dialect = sniffer.sniff(text[:1024])
            except csv.Error:
                # Default to comma if detection fails
                dialect = csv.excel

            reader = csv.reader(io.StringIO(text), dialect)

            # Read header row
            try:
                header_row = next(reader)
            except StopIteration:
                raise ValueError("CSV file is empty")

            # Find column indices
            serial_col, mac_col = self._find_csv_columns(header_row)

            if serial_col is None:
                raise ValueError(
                    f"Could not find Serial Number column. "
                    f"Expected one of: {', '.join(self.SERIAL_COLUMNS)}"
                )

            # Parse data rows
            rows = []
            for row_num, row in enumerate(reader, start=2):
                # Skip empty rows
                if not row or serial_col >= len(row):
                    continue

                serial_value = row[serial_col]
                if serial_value is None or str(serial_value).strip() == "":
                    continue

                # Get MAC address if column exists
                mac_value = None
                if mac_col is not None and mac_col < len(row):
                    mac_value = row[mac_col]
                    if mac_value:
                        mac_value = self._normalize_mac(str(mac_value))

                rows.append(
                    ExcelRow(
                        row_number=row_num,
                        serial_number=str(serial_value).strip(),
                        mac_address=mac_value,
                    )
                )

            logger.info(f"Parsed {len(rows)} rows from CSV file")
            return rows

        except Exception as e:
            if isinstance(e, ValueError):
                raise
            logger.error(f"Failed to parse CSV file: {e}")
            raise ValueError(f"Failed to parse CSV file: {e}")

    def _find_csv_columns(
        self, header_row: list[str]
    ) -> tuple[Optional[int], Optional[int]]:
        """Find the indices of Serial Number and MAC Address columns in CSV.

        Args:
            header_row: List of header strings

        Returns:
            Tuple of (serial_col_index, mac_col_index), None if not found
        """
        serial_col = None
        mac_col = None

        for idx, cell in enumerate(header_row):
            if cell is None:
                continue

            header = str(cell).strip().lower()

            # Check for serial number column
            if header in self.SERIAL_COLUMNS:
                serial_col = idx

            # Check for MAC address column
            elif header in self.MAC_COLUMNS:
                mac_col = idx

        return serial_col, mac_col

    def _parse_excel(self, file_content: bytes) -> list[ExcelRow]:
        """Parse an Excel file.

        Args:
            file_content: Raw bytes of the Excel file

        Returns:
            List of ExcelRow objects

        Raises:
            ValueError: If file format is invalid
        """
        try:
            # Load workbook from bytes
            wb = load_workbook(filename=io.BytesIO(file_content), read_only=True)
            ws = wb.active

            if ws is None:
                raise ValueError("Excel file has no active worksheet")

            # Find column indices
            serial_col, mac_col = self._find_columns(ws)

            if serial_col is None:
                raise ValueError(
                    f"Could not find Serial Number column. "
                    f"Expected one of: {', '.join(self.SERIAL_COLUMNS)}"
                )

            # Parse data rows
            rows = []
            for row_num, row in enumerate(ws.iter_rows(min_row=2), start=2):
                # Skip empty rows
                serial_value = row[serial_col].value if serial_col < len(row) else None
                if serial_value is None or str(serial_value).strip() == "":
                    continue

                # Get MAC address if column exists
                mac_value = None
                if mac_col is not None and mac_col < len(row):
                    mac_value = row[mac_col].value
                    if mac_value:
                        mac_value = self._normalize_mac(str(mac_value))

                rows.append(
                    ExcelRow(
                        row_number=row_num,
                        serial_number=str(serial_value).strip(),
                        mac_address=mac_value,
                    )
                )

            wb.close()

            logger.info(f"Parsed {len(rows)} rows from Excel file")
            return rows

        except Exception as e:
            if isinstance(e, ValueError):
                raise
            logger.error(f"Failed to parse Excel file: {e}")
            raise ValueError(f"Failed to parse Excel file: {e}")

    def validate(self, rows: list[ExcelRow]) -> ValidationResult:
        """Validate parsed Excel rows.

        Args:
            rows: List of ExcelRow to validate

        Returns:
            ValidationResult with any errors
        """
        errors: list[ValidationError] = []
        warnings: list[str] = []

        seen_serials: set[str] = set()

        for row in rows:
            # Check for empty serial number (shouldn't happen after parse, but just in case)
            if not row.serial_number:
                errors.append(
                    ValidationError(
                        row_number=row.row_number,
                        field="serial_number",
                        message="Serial number is required",
                    )
                )
                continue

            # Check for duplicate serial numbers
            serial_upper = row.serial_number.upper()
            if serial_upper in seen_serials:
                errors.append(
                    ValidationError(
                        row_number=row.row_number,
                        field="serial_number",
                        message=f"Duplicate serial number: {row.serial_number}",
                    )
                )
            else:
                seen_serials.add(serial_upper)

            # Validate MAC address format if provided
            if row.mac_address:
                if not self._is_valid_mac(row.mac_address):
                    errors.append(
                        ValidationError(
                            row_number=row.row_number,
                            field="mac_address",
                            message=f"Invalid MAC address format: {row.mac_address}",
                        )
                    )

        # Add warnings for common issues
        if len(rows) > 1000:
            warnings.append(
                f"Large file with {len(rows)} devices. Processing may take a while."
            )

        return ValidationResult(
            is_valid=len(errors) == 0,
            errors=errors,
            warnings=warnings,
        )

    def _find_columns(
        self, ws: Worksheet
    ) -> tuple[Optional[int], Optional[int]]:
        """Find the indices of Serial Number and MAC Address columns.

        Args:
            ws: Excel worksheet

        Returns:
            Tuple of (serial_col_index, mac_col_index), None if not found
        """
        serial_col = None
        mac_col = None

        # Read header row
        header_row = next(ws.iter_rows(min_row=1, max_row=1), None)
        if header_row is None:
            return None, None

        for idx, cell in enumerate(header_row):
            if cell.value is None:
                continue

            header = str(cell.value).strip().lower()

            # Check for serial number column
            if header in self.SERIAL_COLUMNS:
                serial_col = idx

            # Check for MAC address column
            elif header in self.MAC_COLUMNS:
                mac_col = idx

        return serial_col, mac_col

    @staticmethod
    def _normalize_mac(mac: str) -> str:
        """Normalize MAC address to XX:XX:XX:XX:XX:XX format.

        Args:
            mac: MAC address in any format

        Returns:
            Normalized MAC address
        """
        # Remove common separators
        mac = mac.strip().upper()
        mac = mac.replace(":", "").replace("-", "").replace(".", "")

        # Check if we have 12 hex characters
        if len(mac) != 12:
            return mac  # Return as-is, validation will catch it

        # Format as XX:XX:XX:XX:XX:XX
        return ":".join(mac[i : i + 2] for i in range(0, 12, 2))

    @staticmethod
    def _is_valid_mac(mac: str) -> bool:
        """Check if MAC address is valid.

        Args:
            mac: MAC address to validate

        Returns:
            True if valid, False otherwise
        """
        # Check against the pattern
        if MAC_PATTERN.match(mac):
            return True

        # Also accept normalized format
        normalized = mac.replace(":", "").replace("-", "").replace(".", "")
        return len(normalized) == 12 and all(c in "0123456789ABCDEFabcdef" for c in normalized)
