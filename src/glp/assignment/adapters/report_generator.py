"""Report generator adapter.

This adapter implements IReportGenerator to create comprehensive reports
of assignment operations, including phase-by-phase breakdown.
"""

import io
import logging
from datetime import datetime
from typing import Any, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from ..domain.entities import OperationResult
from ..domain.ports import IReportGenerator

logger = logging.getLogger(__name__)


class SimpleReportGenerator(IReportGenerator):
    """Comprehensive report generator.

    Generates JSON reports and Excel files summarizing
    assignment operations with phase-by-phase breakdown.
    """

    def generate(
        self,
        operations: list[OperationResult],
        sync_result: Optional[dict] = None,
        phase_results: Optional[list[dict]] = None,
        workflow_stats: Optional[dict] = None,
    ) -> dict:
        """Generate a JSON report.

        Args:
            operations: List of operation results
            sync_result: Optional sync statistics
            phase_results: Optional phase-by-phase results
            workflow_stats: Optional workflow-level statistics

        Returns:
            Report data structure
        """
        # Calculate statistics
        total = len(operations)
        successful = sum(1 for op in operations if op.success)
        failed = total - successful

        # Count by type
        by_type = {
            "create": {"success": 0, "failed": 0, "devices": 0},
            "application": {"success": 0, "failed": 0, "devices": 0},
            "subscription": {"success": 0, "failed": 0, "devices": 0},
            "tags": {"success": 0, "failed": 0, "devices": 0},
        }

        for op in operations:
            op_type = op.operation_type
            if op_type not in by_type:
                by_type[op_type] = {"success": 0, "failed": 0, "devices": 0}

            device_count = len(op.device_serials or []) or len(op.device_ids or [])
            if op.success:
                by_type[op_type]["success"] += 1
                by_type[op_type]["devices"] += device_count
            else:
                by_type[op_type]["failed"] += 1

        # Build report
        report = {
            "generated_at": datetime.now().isoformat(),
            "summary": {
                "total_operations": total,
                "successful": successful,
                "failed": failed,
                "success_rate": f"{(successful / total * 100):.1f}%" if total > 0 else "N/A",
            },
            "by_operation_type": by_type,
            "sync": sync_result,
            "workflow": workflow_stats,
            "phases": phase_results,
            "errors": [
                {
                    "operation_type": op.operation_type,
                    "devices": op.device_serials,
                    "error": op.error,
                }
                for op in operations
                if not op.success
            ],
        }

        return report

    def generate_excel(
        self,
        operations: list[OperationResult],
        sync_result: Optional[dict] = None,
        phase_results: Optional[list[Any]] = None,
        workflow_stats: Optional[dict] = None,
    ) -> bytes:
        """Generate a comprehensive Excel report.

        Args:
            operations: List of operation results
            sync_result: Optional sync statistics
            phase_results: Optional phase-by-phase results
            workflow_stats: Optional workflow-level statistics

        Returns:
            Excel file bytes
        """
        wb = Workbook()

        # Styling
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        success_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        error_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        warning_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # ========== Summary Sheet ==========
        ws_summary = wb.active
        ws_summary.title = "Summary"

        # Title
        ws_summary["A1"] = "Device Assignment Report"
        ws_summary["A1"].font = Font(bold=True, size=16)
        ws_summary.merge_cells("A1:D1")

        # Generation info
        ws_summary["A3"] = "Generated At:"
        ws_summary["B3"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        # Overall statistics
        total = len(operations)
        successful = sum(1 for op in operations if op.success)

        ws_summary["A5"] = "Overall Statistics"
        ws_summary["A5"].font = Font(bold=True, size=12)

        stats_data = [
            ("Total Operations:", total),
            ("Successful:", successful),
            ("Failed:", total - successful),
            ("Success Rate:", f"{(successful / total * 100):.1f}%" if total > 0 else "N/A"),
        ]

        for i, (label, value) in enumerate(stats_data, start=6):
            ws_summary[f"A{i}"] = label
            ws_summary[f"B{i}"] = value

        # Breakdown by operation type
        ws_summary["A11"] = "Breakdown by Operation Type"
        ws_summary["A11"].font = Font(bold=True, size=12)

        breakdown_headers = ["Operation Type", "Success", "Failed", "Devices Affected"]
        for col, header in enumerate(breakdown_headers, 1):
            cell = ws_summary.cell(row=12, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = thin_border

        by_type = {}
        for op in operations:
            op_type = op.operation_type
            if op_type not in by_type:
                by_type[op_type] = {"success": 0, "failed": 0, "devices": 0}
            device_count = len(op.device_serials or []) or len(op.device_ids or [])
            if op.success:
                by_type[op_type]["success"] += 1
                by_type[op_type]["devices"] += device_count
            else:
                by_type[op_type]["failed"] += 1

        row = 13
        for op_type, stats in by_type.items():
            ws_summary.cell(row=row, column=1, value=op_type.title()).border = thin_border
            ws_summary.cell(row=row, column=2, value=stats["success"]).border = thin_border
            ws_summary.cell(row=row, column=3, value=stats["failed"]).border = thin_border
            ws_summary.cell(row=row, column=4, value=stats["devices"]).border = thin_border
            row += 1

        # Workflow stats if available
        if workflow_stats:
            row += 2
            ws_summary.cell(row=row, column=1, value="Workflow Statistics").font = Font(bold=True, size=12)
            row += 1

            workflow_items = [
                ("Devices Created:", workflow_stats.get("devices_created", 0)),
                ("Applications Assigned:", workflow_stats.get("applications_assigned", 0)),
                ("Subscriptions Assigned:", workflow_stats.get("subscriptions_assigned", 0)),
                ("Tags Updated:", workflow_stats.get("tags_updated", 0)),
                ("Total Duration:", f"{workflow_stats.get('total_duration_seconds', 0):.1f}s"),
            ]

            for label, value in workflow_items:
                ws_summary.cell(row=row, column=1, value=label)
                ws_summary.cell(row=row, column=2, value=value)
                row += 1

            # New devices info
            if workflow_stats.get("new_devices_added"):
                row += 1
                ws_summary.cell(row=row, column=1, value="New Devices Added:").font = Font(bold=True)
                row += 1
                for serial in workflow_stats.get("new_devices_added", []):
                    cell = ws_summary.cell(row=row, column=1, value=serial)
                    cell.fill = success_fill
                    row += 1

            if workflow_stats.get("new_devices_failed"):
                row += 1
                ws_summary.cell(row=row, column=1, value="New Devices Failed:").font = Font(bold=True)
                row += 1
                for serial in workflow_stats.get("new_devices_failed", []):
                    cell = ws_summary.cell(row=row, column=1, value=serial)
                    cell.fill = error_fill
                    row += 1

        # Sync results if available
        if sync_result:
            row += 2
            ws_summary.cell(row=row, column=1, value="Sync Results").font = Font(bold=True, size=12)
            row += 1

            sync_items = [
                ("Devices Synced:", sync_result.get("devices_synced", sync_result.get("records_fetched", 0))),
                ("Subscriptions Synced:", sync_result.get("subscriptions_synced", 0)),
            ]

            for label, value in sync_items:
                ws_summary.cell(row=row, column=1, value=label)
                ws_summary.cell(row=row, column=2, value=value)
                row += 1

        # Column widths
        ws_summary.column_dimensions["A"].width = 25
        ws_summary.column_dimensions["B"].width = 20
        ws_summary.column_dimensions["C"].width = 15
        ws_summary.column_dimensions["D"].width = 18

        # ========== Phases Sheet (if available) ==========
        if phase_results:
            ws_phases = wb.create_sheet("Phases")

            ws_phases["A1"] = "Workflow Phases"
            ws_phases["A1"].font = Font(bold=True, size=14)

            phase_headers = ["Phase", "Status", "Devices Processed", "Errors", "Duration (s)"]
            for col, header in enumerate(phase_headers, 1):
                cell = ws_phases.cell(row=3, column=col, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.border = thin_border

            for row_num, phase in enumerate(phase_results, 4):
                # Handle both dict and PhaseResult objects
                if hasattr(phase, "phase_name"):
                    phase_name = phase.phase_name
                    success = phase.success
                    devices = phase.devices_processed
                    errors = phase.errors
                    duration = phase.duration_seconds
                else:
                    phase_name = phase.get("phase_name", "Unknown")
                    success = phase.get("success", False)
                    devices = phase.get("devices_processed", 0)
                    errors = phase.get("errors", 0)
                    duration = phase.get("duration_seconds", 0)

                ws_phases.cell(row=row_num, column=1, value=phase_name).border = thin_border

                status_cell = ws_phases.cell(
                    row=row_num, column=2, value="Success" if success else "Failed"
                )
                status_cell.fill = success_fill if success else error_fill
                status_cell.border = thin_border

                ws_phases.cell(row=row_num, column=3, value=devices).border = thin_border
                ws_phases.cell(row=row_num, column=4, value=errors).border = thin_border
                ws_phases.cell(row=row_num, column=5, value=f"{duration:.2f}").border = thin_border

            # Column widths
            ws_phases.column_dimensions["A"].width = 25
            ws_phases.column_dimensions["B"].width = 12
            ws_phases.column_dimensions["C"].width = 18
            ws_phases.column_dimensions["D"].width = 10
            ws_phases.column_dimensions["E"].width = 14

        # ========== Operations Sheet ==========
        ws_ops = wb.create_sheet("Operations")

        # Headers
        headers = ["Operation Type", "Status", "Device Serials", "Device Count", "Error"]
        for col, header in enumerate(headers, 1):
            cell = ws_ops.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border

        # Data rows
        for row_num, op in enumerate(operations, 2):
            ws_ops.cell(row=row_num, column=1, value=op.operation_type.title()).border = thin_border

            status_cell = ws_ops.cell(
                row=row_num, column=2, value="Success" if op.success else "Failed"
            )
            status_cell.fill = success_fill if op.success else error_fill
            status_cell.border = thin_border

            # Join device serials/IDs
            devices = op.device_serials or [str(d) for d in (op.device_ids or [])]
            device_str = ", ".join(devices[:10])  # Limit to first 10
            if len(devices) > 10:
                device_str += f" (+{len(devices) - 10} more)"
            ws_ops.cell(row=row_num, column=3, value=device_str).border = thin_border

            ws_ops.cell(row=row_num, column=4, value=len(devices)).border = thin_border

            if op.error:
                error_cell = ws_ops.cell(row=row_num, column=5, value=op.error)
                error_cell.border = thin_border

        # Column widths
        ws_ops.column_dimensions["A"].width = 18
        ws_ops.column_dimensions["B"].width = 12
        ws_ops.column_dimensions["C"].width = 60
        ws_ops.column_dimensions["D"].width = 14
        ws_ops.column_dimensions["E"].width = 50

        # ========== Errors Sheet (if any) ==========
        errors = [op for op in operations if not op.success]
        if errors:
            ws_errors = wb.create_sheet("Errors")

            ws_errors["A1"] = "Error Details"
            ws_errors["A1"].font = Font(bold=True, size=14, color="FF0000")

            # Headers
            error_headers = ["#", "Operation Type", "Device Serials", "Error Message"]
            for col, header in enumerate(error_headers, 1):
                cell = ws_errors.cell(row=3, column=col, value=header)
                cell.fill = error_fill
                cell.font = Font(bold=True)
                cell.border = thin_border

            # Error rows
            for row_num, op in enumerate(errors, 4):
                ws_errors.cell(row=row_num, column=1, value=row_num - 3).border = thin_border
                ws_errors.cell(row=row_num, column=2, value=op.operation_type.title()).border = thin_border

                devices = op.device_serials or [str(d) for d in (op.device_ids or [])]
                ws_errors.cell(row=row_num, column=3, value=", ".join(devices)).border = thin_border
                ws_errors.cell(row=row_num, column=4, value=op.error or "Unknown error").border = thin_border

            # Column widths
            ws_errors.column_dimensions["A"].width = 5
            ws_errors.column_dimensions["B"].width = 18
            ws_errors.column_dimensions["C"].width = 50
            ws_errors.column_dimensions["D"].width = 70

        # ========== Devices Sheet (detailed) ==========
        ws_devices = wb.create_sheet("Devices")

        ws_devices["A1"] = "Device Details"
        ws_devices["A1"].font = Font(bold=True, size=14)

        device_headers = ["Serial Number", "Status", "Application", "Subscription", "Tags", "Errors"]
        for col, header in enumerate(device_headers, 1):
            cell = ws_devices.cell(row=3, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = thin_border

        # Aggregate device info from operations
        device_info: dict[str, dict] = {}
        for op in operations:
            for serial in (op.device_serials or []):
                if serial not in device_info:
                    device_info[serial] = {
                        "status": "pending",
                        "application": False,
                        "subscription": False,
                        "tags": False,
                        "errors": [],
                    }

                if op.success:
                    if op.operation_type == "create":
                        device_info[serial]["status"] = "created"
                    elif op.operation_type == "application":
                        device_info[serial]["application"] = True
                    elif op.operation_type == "subscription":
                        device_info[serial]["subscription"] = True
                    elif op.operation_type == "tags":
                        device_info[serial]["tags"] = True
                else:
                    device_info[serial]["errors"].append(f"{op.operation_type}: {op.error}")

        row = 4
        for serial, info in sorted(device_info.items()):
            ws_devices.cell(row=row, column=1, value=serial).border = thin_border

            # Determine overall status
            if info["errors"]:
                status = "Error"
                status_fill = error_fill
            elif info["status"] == "created":
                status = "Created"
                status_fill = success_fill
            elif info["application"] or info["subscription"] or info["tags"]:
                status = "Updated"
                status_fill = success_fill
            else:
                status = "Pending"
                status_fill = warning_fill

            status_cell = ws_devices.cell(row=row, column=2, value=status)
            status_cell.fill = status_fill
            status_cell.border = thin_border

            app_cell = ws_devices.cell(row=row, column=3, value="Yes" if info["application"] else "No")
            app_cell.fill = success_fill if info["application"] else PatternFill()
            app_cell.border = thin_border

            sub_cell = ws_devices.cell(row=row, column=4, value="Yes" if info["subscription"] else "No")
            sub_cell.fill = success_fill if info["subscription"] else PatternFill()
            sub_cell.border = thin_border

            tag_cell = ws_devices.cell(row=row, column=5, value="Yes" if info["tags"] else "No")
            tag_cell.fill = success_fill if info["tags"] else PatternFill()
            tag_cell.border = thin_border

            ws_devices.cell(row=row, column=6, value="; ".join(info["errors"]) if info["errors"] else "").border = thin_border

            row += 1

        # Column widths
        ws_devices.column_dimensions["A"].width = 20
        ws_devices.column_dimensions["B"].width = 12
        ws_devices.column_dimensions["C"].width = 12
        ws_devices.column_dimensions["D"].width = 14
        ws_devices.column_dimensions["E"].width = 10
        ws_devices.column_dimensions["F"].width = 50

        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        return output.getvalue()
