"""Assignment template generator.

Generates sample CSV/Excel templates for the device assignment workflow
with clear instructions and example data.
"""

import csv
import io
import logging
from typing import Any

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from .generator import BaseReportGenerator

logger = logging.getLogger(__name__)


class AssignmentTemplateGenerator(BaseReportGenerator):
    """Generate sample templates for device assignment workflow."""

    # Template columns with descriptions
    TEMPLATE_COLUMNS = [
        ("serial_number", "Serial Number", "Device serial number (required)", True),
        ("mac_address", "MAC Address", "Device MAC address in XX:XX:XX:XX:XX:XX format (optional)", False),
        ("device_type", "Device Type", "AP, SWITCH, GATEWAY, or IAP (optional - auto-detected)", False),
        ("subscription_key", "Subscription Key", "Subscription key to assign (optional)", False),
        ("application_id", "Application ID", "Application ID for region assignment (optional)", False),
        ("tags", "Tags", "Tags as key:value pairs separated by semicolons (optional)", False),
    ]

    # Example data for the template
    EXAMPLE_DATA = [
        {
            "serial_number": "CN12345ABC",
            "mac_address": "00:11:22:33:44:55",
            "device_type": "AP",
            "subscription_key": "SUB-KEY-001",
            "application_id": "app-us-west-001",
            "tags": "environment:production;team:networking",
        },
        {
            "serial_number": "CN67890DEF",
            "mac_address": "66:77:88:99:AA:BB",
            "device_type": "SWITCH",
            "subscription_key": "SUB-KEY-002",
            "application_id": "app-us-east-001",
            "tags": "environment:staging;team:infrastructure",
        },
        {
            "serial_number": "CN11223GHI",
            "mac_address": "CC:DD:EE:FF:00:11",
            "device_type": "GATEWAY",
            "subscription_key": "",
            "application_id": "",
            "tags": "location:datacenter-1",
        },
    ]

    def generate_excel(
        self,
        data: dict[str, Any] | None = None,
        filters: dict[str, Any] | None = None,
    ) -> bytes:
        """Generate Excel template with instructions and example data.

        The template includes:
        1. Instructions sheet with detailed guidance
        2. Template sheet with column headers and examples
        3. Data validation where applicable
        """
        wb = Workbook()

        # Sheet 1: Instructions
        ws_instructions = wb.active
        ws_instructions.title = "Instructions"
        self._create_instructions_sheet(ws_instructions)

        # Sheet 2: Template
        ws_template = wb.create_sheet("Device Template")
        self._create_template_sheet(ws_template)

        return self._workbook_to_bytes(wb)

    def generate_csv(
        self,
        data: dict[str, Any] | None = None,
        filters: dict[str, Any] | None = None,
    ) -> str:
        """Generate simple CSV template with headers and examples."""
        output = io.StringIO()

        fieldnames = [col[0] for col in self.TEMPLATE_COLUMNS]
        writer = csv.DictWriter(output, fieldnames=fieldnames, quoting=csv.QUOTE_MINIMAL)
        writer.writeheader()

        # Write example rows
        for example in self.EXAMPLE_DATA:
            writer.writerow(example)

        return output.getvalue()

    def generate_json(
        self,
        data: dict[str, Any] | None = None,
        filters: dict[str, Any] | None = None,
    ) -> str:
        """Generate JSON template with structure and examples.

        Returns a JSON object containing:
        - Template schema with column definitions
        - Example data
        - Instructions and formatting rules
        """
        import json

        # Build column schema
        columns = []
        for col_name, display_name, description, required in self.TEMPLATE_COLUMNS:
            columns.append({
                "field": col_name,
                "label": display_name,
                "description": description,
                "required": required,
            })

        template = {
            "template_version": "1.0",
            "description": "Device assignment template for bulk subscription and tag assignments",
            "instructions": {
                "overview": "This template allows you to bulk assign subscriptions, applications, and tags to devices.",
                "usage": [
                    "Only the Serial Number field is required - all other fields are optional.",
                    "Delete the example data and add your device information.",
                    "Devices not found in the database will be flagged for review.",
                    "You can assign the same subscription to multiple devices.",
                    "Existing assignments will be updated, not duplicated.",
                    "Leave fields empty to skip assignment for that attribute.",
                ],
                "tags_format": {
                    "description": "Tags should be formatted as key:value pairs separated by semicolons",
                    "example": "environment:production;team:networking;owner:john",
                    "rules": [
                        "Tag keys and values cannot contain colons or semicolons",
                        "Multiple tags are separated by semicolons",
                        "Each tag is a key:value pair",
                    ],
                },
                "supported_device_types": [
                    {"code": "AP", "name": "Access Point"},
                    {"code": "IAP", "name": "Instant Access Point"},
                    {"code": "SWITCH", "name": "Network Switch"},
                    {"code": "GATEWAY", "name": "Gateway Device"},
                ],
            },
            "columns": columns,
            "examples": self.EXAMPLE_DATA,
        }

        return json.dumps(template, indent=2)

    def _create_instructions_sheet(self, ws) -> None:
        """Create instructions sheet with detailed guidance."""
        # Styling
        title_font = Font(bold=True, size=16, color="01A982")
        header_font = Font(bold=True, size=12, color="305496")
        normal_font = Font(size=11)
        note_font = Font(size=10, italic=True, color="595959")

        row = 1

        # Title
        ws.cell(row=row, column=1, value="Device Assignment Template Instructions")
        ws.cell(row=row, column=1).font = title_font
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        row += 2

        # Overview
        ws.cell(row=row, column=1, value="Overview")
        ws.cell(row=row, column=1).font = header_font
        row += 1

        instructions = [
            "This template allows you to bulk assign subscriptions, applications, and tags to devices.",
            "Fill in the 'Device Template' sheet with your device information.",
            "Only the Serial Number column is required - all other columns are optional.",
            "",
        ]

        for text in instructions:
            ws.cell(row=row, column=1, value=text)
            ws.cell(row=row, column=1).font = normal_font
            row += 1

        row += 1

        # Column descriptions
        ws.cell(row=row, column=1, value="Column Descriptions")
        ws.cell(row=row, column=1).font = header_font
        row += 2

        # Headers for column table
        headers = ["Column", "Required", "Description"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        row += 1

        for col_name, display_name, description, required in self.TEMPLATE_COLUMNS:
            ws.cell(row=row, column=1, value=display_name)
            ws.cell(row=row, column=2, value="Yes" if required else "No")
            ws.cell(row=row, column=3, value=description)
            row += 1

        row += 2

        # Tags format
        ws.cell(row=row, column=1, value="Tags Format")
        ws.cell(row=row, column=1).font = header_font
        row += 1

        tag_instructions = [
            "Tags should be formatted as key:value pairs separated by semicolons.",
            'Example: "environment:production;team:networking;owner:john"',
            "Tag keys and values cannot contain colons or semicolons.",
            "",
        ]

        for text in tag_instructions:
            ws.cell(row=row, column=1, value=text)
            ws.cell(row=row, column=1).font = normal_font
            row += 1

        row += 1

        # Device Types
        ws.cell(row=row, column=1, value="Supported Device Types")
        ws.cell(row=row, column=1).font = header_font
        row += 1

        device_types = [
            "AP - Access Point",
            "IAP - Instant Access Point",
            "SWITCH - Network Switch",
            "GATEWAY - Gateway Device",
        ]

        for dtype in device_types:
            ws.cell(row=row, column=1, value=f"  {dtype}")
            ws.cell(row=row, column=1).font = normal_font
            row += 1

        row += 2

        # Notes
        ws.cell(row=row, column=1, value="Important Notes")
        ws.cell(row=row, column=1).font = header_font
        row += 1

        notes = [
            "1. Delete the example rows before importing your data.",
            "2. Devices not found in the database will be flagged for review.",
            "3. You can assign the same subscription to multiple devices.",
            "4. Existing assignments will be updated, not duplicated.",
            "5. Leave fields empty to skip assignment for that attribute.",
        ]

        for note in notes:
            ws.cell(row=row, column=1, value=note)
            ws.cell(row=row, column=1).font = note_font
            row += 1

        # Column widths
        ws.column_dimensions["A"].width = 50
        ws.column_dimensions["B"].width = 12
        ws.column_dimensions["C"].width = 60

    def _create_template_sheet(self, ws) -> None:
        """Create the data template sheet with headers and examples."""
        # Styling
        header_fill = PatternFill(start_color="01A982", end_color="01A982", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        example_fill = PatternFill(start_color="E6F9F3", end_color="E6F9F3", fill_type="solid")
        border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # Headers with comments
        for col, (col_name, display_name, description, required) in enumerate(self.TEMPLATE_COLUMNS, 1):
            cell = ws.cell(row=1, column=col, value=display_name)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = Alignment(horizontal="center")

            # Add comment with description
            comment = Comment(description, "HPE GreenLake")
            comment.width = 300
            comment.height = 50
            cell.comment = comment

        # Example data rows
        for row_idx, example in enumerate(self.EXAMPLE_DATA, 2):
            for col_idx, (col_name, _, _, _) in enumerate(self.TEMPLATE_COLUMNS, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=example.get(col_name, ""))
                cell.fill = example_fill
                cell.border = border
                cell.font = Font(italic=True, color="595959")

        # Add data validation for device_type column
        device_type_col = 3  # Column C
        dv = DataValidation(
            type="list",
            formula1='"AP,IAP,SWITCH,GATEWAY"',
            allow_blank=True,
        )
        dv.error = "Please select a valid device type"
        dv.errorTitle = "Invalid Device Type"
        dv.prompt = "Select a device type or leave blank for auto-detection"
        dv.promptTitle = "Device Type"
        ws.add_data_validation(dv)
        dv.add(f"{get_column_letter(device_type_col)}2:{get_column_letter(device_type_col)}1000")

        # Column widths
        column_widths = [20, 20, 15, 25, 25, 45]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = width

        # Freeze header row
        ws.freeze_panes = "A2"

        # Add note at the bottom
        note_row = len(self.EXAMPLE_DATA) + 3
        ws.cell(
            row=note_row,
            column=1,
            value="Note: Delete the example rows above and add your device data. Only Serial Number is required.",
        )
        ws.cell(row=note_row, column=1).font = Font(italic=True, color="595959", size=9)
        ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=6)
