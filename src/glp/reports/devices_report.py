"""Device inventory report generator.

Generates comprehensive device inventory reports with:
- Summary statistics
- Full device listing with all fields
- Insights on unassigned and expiring devices
- Aruba Central integration status
"""

import json
import logging
from typing import Any

from openpyxl.styles import Font

from .generator import BaseReportGenerator

logger = logging.getLogger(__name__)


class DevicesReportGenerator(BaseReportGenerator):
    """Generate detailed device inventory reports."""

    # Column definitions for the device list
    DEVICE_COLUMNS = [
        ("serial_number", "Serial Number"),
        ("mac_address", "MAC Address"),
        ("device_type", "Device Type"),
        ("model", "Model"),
        ("region", "Region"),
        ("device_name", "Device Name"),
        ("assigned_state", "Assignment Status"),
        ("location_city", "City"),
        ("location_country", "Country"),
        ("subscription_key", "Subscription Key"),
        ("subscription_type", "Subscription Type"),
        ("subscription_end", "Subscription End"),
        ("tags", "Tags"),
        ("central_status", "Central Status"),
        ("central_ipv4", "Central IP"),
        ("central_site_name", "Central Site"),
        ("central_software_version", "Software Version"),
        ("updated_at", "Last Updated"),
    ]

    def generate_excel(
        self,
        data: dict[str, Any],
        filters: dict[str, Any] | None = None,
    ) -> bytes:
        """Generate multi-sheet device inventory Excel report.

        Sheets:
        1. Summary - Statistics and breakdown
        2. Device List - Full device inventory
        3. Insights - Devices needing attention
        """
        wb = self.create_workbook()

        # Extract items
        items = data.get("items", [])
        total = data.get("total", len(items))

        # Sheet 1: Summary
        ws_summary = wb.active
        ws_summary.title = "Summary"
        self._create_summary_sheet(ws_summary, items, total, filters)

        # Sheet 2: Device List
        ws_devices = wb.create_sheet("Device List")
        self._create_device_list_sheet(ws_devices, items, filters)

        # Sheet 3: Insights
        ws_insights = wb.create_sheet("Insights")
        self._create_insights_sheet(ws_insights, items)

        return self._workbook_to_bytes(wb)

    def generate_csv(
        self,
        data: dict[str, Any],
        filters: dict[str, Any] | None = None,
    ) -> str:
        """Generate CSV with device inventory."""
        items = data.get("items", [])

        # Flatten the data for CSV
        rows = []
        for item in items:
            row = {}
            for field, _ in self.DEVICE_COLUMNS:
                value = item.get(field, "")
                # Handle nested dict (tags)
                if field == "tags" and isinstance(value, dict):
                    value = json.dumps(value) if value else ""
                row[field] = value
            rows.append(row)

        fieldnames = [field for field, _ in self.DEVICE_COLUMNS]
        return self._dict_to_csv(rows, fieldnames)

    def _create_summary_sheet(
        self,
        ws,
        items: list[dict],
        total: int,
        filters: dict[str, Any] | None,
    ) -> None:
        """Create summary sheet with statistics."""
        row = self.add_report_header(
            ws,
            "Device Inventory Report",
            f"{total:,} Devices",
            filters,
        )

        # Quick Stats KPIs
        ws.cell(row=row, column=1, value="Quick Statistics")
        ws.cell(row=row, column=1).font = self.styles.get_subtitle_font()
        row += 2

        # Calculate statistics
        assigned = sum(1 for d in items if d.get("assigned_state") == "ASSIGNED_TO_SERVICE")
        in_central = sum(1 for d in items if d.get("in_central"))
        online = sum(1 for d in items if d.get("central_status") == "ONLINE")
        with_subscription = sum(1 for d in items if d.get("subscription_key"))

        kpi_data = [
            ("Total Devices", len(items), None),
            ("Assigned", assigned, f"{(assigned / max(len(items), 1) * 100):.1f}%"),
            ("In Aruba Central", in_central, f"{online} online"),
            ("With Subscription", with_subscription, None),
        ]

        col = 1
        kpi_row = row
        for label, value, subtitle in kpi_data:
            self.add_kpi_card(ws, kpi_row, col, label, value, subtitle)
            col += 2

        row = kpi_row + 6

        # Device Type Breakdown
        ws.cell(row=row, column=1, value="Breakdown by Device Type")
        ws.cell(row=row, column=1).font = self.styles.get_subtitle_font()
        row += 2

        type_counts: dict[str, int] = {}
        for item in items:
            dtype = item.get("device_type", "Unknown") or "Unknown"
            type_counts[dtype] = type_counts.get(dtype, 0) + 1

        headers = ["Device Type", "Count", "Percentage"]
        self.add_table_headers(ws, headers, row)
        row += 1

        for dtype, count in sorted(type_counts.items(), key=lambda x: -x[1]):
            pct = f"{(count / max(len(items), 1) * 100):.1f}%"
            self.add_data_row(ws, [dtype, count, pct], row, alternate=(row % 2 == 0))
            row += 1

        row += 2

        # Region Breakdown
        ws.cell(row=row, column=1, value="Breakdown by Region")
        ws.cell(row=row, column=1).font = self.styles.get_subtitle_font()
        row += 2

        region_counts: dict[str, int] = {}
        for item in items:
            region = item.get("region", "Unknown") or "Unknown"
            region_counts[region] = region_counts.get(region, 0) + 1

        self.add_table_headers(ws, ["Region", "Count", "Percentage"], row)
        row += 1

        for region, count in sorted(region_counts.items(), key=lambda x: -x[1]):
            pct = f"{(count / max(len(items), 1) * 100):.1f}%"
            self.add_data_row(ws, [region, count, pct], row, alternate=(row % 2 == 0))
            row += 1

        self.auto_fit_columns(ws)
        self.freeze_panes(ws, row=8)

    def _create_device_list_sheet(
        self,
        ws,
        items: list[dict],
        filters: dict[str, Any] | None,
    ) -> None:
        """Create the main device list sheet."""
        row = 1

        # Headers
        headers = [label for _, label in self.DEVICE_COLUMNS]
        self.add_table_headers(ws, headers, row)
        row += 1

        # Data rows
        for item in items:
            row_data = []
            for field, _ in self.DEVICE_COLUMNS:
                value = item.get(field, "")

                # Handle special fields
                if field == "tags" and isinstance(value, dict):
                    # Format tags as key:value pairs
                    value = "; ".join(f"{k}:{v}" for k, v in value.items()) if value else ""
                elif field == "assigned_state":
                    value = "Assigned" if value == "ASSIGNED_TO_SERVICE" else "Unassigned" if value == "UNASSIGNED" else value
                elif field == "subscription_type" and value:
                    value = value.replace("CENTRAL_", "")

                row_data.append(value)

            self.add_data_row(ws, row_data, row, alternate=(row % 2 == 0))

            # Color code assignment status
            assigned_col = 7  # assigned_state column
            cell = ws.cell(row=row, column=assigned_col)
            if item.get("assigned_state") == "ASSIGNED_TO_SERVICE":
                cell.fill = self.styles.get_success_fill()
            elif item.get("assigned_state") == "UNASSIGNED":
                cell.fill = self.styles.get_warning_fill()

            # Color code Central status
            central_col = 14  # central_status column
            central_status = item.get("central_status")
            if central_status:
                cell = ws.cell(row=row, column=central_col)
                cell.fill = self.styles.get_status_fill(central_status)

            row += 1

        self.auto_fit_columns(ws, max_width=40)
        self.freeze_panes(ws, row=2)

    def _create_insights_sheet(self, ws, items: list[dict]) -> None:
        """Create insights sheet highlighting devices needing attention."""
        row = self.add_report_header(
            ws,
            "Device Insights",
            "Devices Requiring Attention",
        )

        # Unassigned devices
        ws.cell(row=row, column=1, value="Unassigned Devices")
        ws.cell(row=row, column=1).font = self.styles.get_subtitle_font()
        row += 2

        unassigned = [d for d in items if d.get("assigned_state") == "UNASSIGNED"]

        if unassigned:
            headers = ["Serial Number", "Device Type", "Model", "Region"]
            self.add_table_headers(ws, headers, row)
            row += 1

            for device in unassigned[:50]:  # Limit to 50
                self.add_data_row(
                    ws,
                    [
                        device.get("serial_number"),
                        device.get("device_type"),
                        device.get("model"),
                        device.get("region"),
                    ],
                    row,
                    alternate=(row % 2 == 0),
                )
                row += 1

            if len(unassigned) > 50:
                ws.cell(row=row, column=1, value=f"... and {len(unassigned) - 50} more")
                ws.cell(row=row, column=1).font = Font(italic=True)
                row += 1
        else:
            ws.cell(row=row, column=1, value="All devices are assigned.")
            ws.cell(row=row, column=1).fill = self.styles.get_success_fill()
            row += 1

        row += 2

        # Devices without subscription
        ws.cell(row=row, column=1, value="Devices Without Subscription")
        ws.cell(row=row, column=1).font = self.styles.get_subtitle_font()
        row += 2

        no_subscription = [d for d in items if not d.get("subscription_key")]

        if no_subscription:
            headers = ["Serial Number", "Device Type", "Model", "Assignment Status"]
            self.add_table_headers(ws, headers, row)
            row += 1

            for device in no_subscription[:50]:
                status = "Assigned" if device.get("assigned_state") == "ASSIGNED_TO_SERVICE" else "Unassigned"
                self.add_data_row(
                    ws,
                    [
                        device.get("serial_number"),
                        device.get("device_type"),
                        device.get("model"),
                        status,
                    ],
                    row,
                    alternate=(row % 2 == 0),
                )
                row += 1

            if len(no_subscription) > 50:
                ws.cell(row=row, column=1, value=f"... and {len(no_subscription) - 50} more")
                ws.cell(row=row, column=1).font = Font(italic=True)
                row += 1
        else:
            ws.cell(row=row, column=1, value="All devices have subscriptions.")
            ws.cell(row=row, column=1).fill = self.styles.get_success_fill()
            row += 1

        row += 2

        # Offline devices in Aruba Central
        ws.cell(row=row, column=1, value="Offline Devices in Aruba Central")
        ws.cell(row=row, column=1).font = self.styles.get_subtitle_font()
        row += 2

        offline = [d for d in items if d.get("in_central") and d.get("central_status") == "OFFLINE"]

        if offline:
            headers = ["Serial Number", "Device Name", "Central Site", "Last Seen"]
            self.add_table_headers(ws, headers, row)
            row += 1

            for device in offline[:50]:
                self.add_data_row(
                    ws,
                    [
                        device.get("serial_number"),
                        device.get("central_device_name"),
                        device.get("central_site_name"),
                        device.get("central_last_seen_at"),
                    ],
                    row,
                    alternate=(row % 2 == 0),
                )
                # Highlight as error
                for col in range(1, 5):
                    ws.cell(row=row, column=col).fill = self.styles.get_error_fill()
                row += 1

            if len(offline) > 50:
                ws.cell(row=row, column=1, value=f"... and {len(offline) - 50} more")
                ws.cell(row=row, column=1).font = Font(italic=True)
                row += 1
        else:
            ws.cell(row=row, column=1, value="All Central devices are online.")
            ws.cell(row=row, column=1).fill = self.styles.get_success_fill()
            row += 1

        self.auto_fit_columns(ws)
        self.freeze_panes(ws, row=6)
