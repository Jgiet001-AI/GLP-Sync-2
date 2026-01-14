"""Client network report generator.

Generates comprehensive network client reports with:
- Client health and connectivity status
- Site-level statistics with visual charts
- Network segmentation analysis
- Connection details
- Visual charts (pie charts, bar charts)
"""

import logging
from typing import Any

from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.styles import Font

from .generator import BaseReportGenerator

logger = logging.getLogger(__name__)

# Chart colors - HPE brand colors
CHART_COLORS = [
    "01A982",  # HPE Green
    "7630EA",  # Purple
    "FF8300",  # Orange
    "00739D",  # Blue
    "C6C9CA",  # Gray
    "E5004C",  # Red
    "FEC422",  # Yellow
]


class ClientsReportGenerator(BaseReportGenerator):
    """Generate detailed network client reports."""

    CLIENT_COLUMNS = [
        ("mac", "MAC Address"),
        ("name", "Client Name"),
        ("site_name", "Site"),
        ("health", "Health"),
        ("status", "Status"),
        ("type", "Type"),
        ("ipv4", "IPv4"),
        ("ipv6", "IPv6"),
        ("network", "Network/SSID"),
        ("vlan_id", "VLAN"),
        ("port", "Port"),
        ("connected_to", "Connected To"),
        ("connected_since", "Connected Since"),
        ("last_seen_at", "Last Seen"),
        ("authentication", "Authentication"),
        ("key_management", "Key Management"),
    ]

    def generate_excel(
        self,
        data: dict[str, Any],
        filters: dict[str, Any] | None = None,
    ) -> bytes:
        """Generate multi-sheet client network Excel report.

        Sheets:
        1. Dashboard - Visual charts and executive overview
        2. Summary - Health and connectivity overview with KPIs
        3. Client List - Full client inventory
        4. Site Statistics - Per-site breakdown with rich details
        5. Network Analysis - VLAN and SSID distribution
        """
        wb = self.create_workbook()

        items = data.get("items", [])
        total = data.get("total", len(items))
        summary = data.get("summary", {})

        # Calculate summary if not provided
        if not summary:
            summary = self._calculate_summary(items)

        # Sheet 1: Dashboard with Charts (new!)
        ws_dashboard = wb.active
        ws_dashboard.title = "Dashboard"
        self._create_dashboard_sheet(ws_dashboard, items, total, summary, filters)

        # Sheet 2: Summary
        ws_summary = wb.create_sheet("Summary")
        self._create_summary_sheet(ws_summary, items, total, summary, filters)

        # Sheet 3: Client List
        ws_clients = wb.create_sheet("Client List")
        self._create_client_list_sheet(ws_clients, items)

        # Sheet 4: Site Statistics (enhanced)
        ws_sites = wb.create_sheet("Site Statistics")
        self._create_site_stats_sheet(ws_sites, items)

        # Sheet 5: Network Analysis
        ws_network = wb.create_sheet("Network Analysis")
        self._create_network_analysis_sheet(ws_network, items)

        return self._workbook_to_bytes(wb)

    def _calculate_summary(self, items: list[dict]) -> dict[str, Any]:
        """Calculate summary statistics from items."""
        return {
            "total_clients": len(items),
            "connected": sum(1 for c in items if c.get("status") == "Connected"),
            "wired": sum(1 for c in items if c.get("type") == "Wired"),
            "wireless": sum(1 for c in items if c.get("type") == "Wireless"),
            "health_good": sum(1 for c in items if c.get("health") == "Good"),
            "health_fair": sum(1 for c in items if c.get("health") == "Fair"),
            "health_poor": sum(1 for c in items if c.get("health") == "Poor"),
            "health_unknown": sum(1 for c in items if c.get("health") not in ("Good", "Fair", "Poor")),
        }

    def _create_dashboard_sheet(
        self,
        ws,
        items: list[dict],
        total: int,
        summary: dict[str, Any],
        filters: dict[str, Any] | None,
    ) -> None:
        """Create dashboard sheet with visual charts."""
        row = self.add_report_header(
            ws,
            "Network Clients Dashboard",
            f"{total:,} Clients - Executive Overview",
            filters,
        )

        # KPI Cards Row
        ws.cell(row=row, column=1, value="Key Metrics")
        ws.cell(row=row, column=1).font = self.styles.get_subtitle_font()
        row += 2

        total_clients = summary.get("total_clients", len(items)) or 0
        connected = summary.get("connected", 0)
        wireless = summary.get("wireless", 0)
        wired = summary.get("wired", 0)
        health_good = summary.get("health_good", 0)

        # Safe division for percentages when no data
        def pct(value: int, total: int) -> str:
            if total == 0:
                return "0.0%"
            return f"{(value / total * 100):.1f}%"

        site_count = self._count_sites(items)
        kpi_data = [
            ("Total Clients", total_clients, f"In {site_count} site{'s' if site_count != 1 else ''}"),
            ("Connected", connected, f"{pct(connected, total_clients)} online"),
            ("Wireless", wireless, pct(wireless, total_clients)),
            ("Wired", wired, pct(wired, total_clients)),
            ("Healthy", health_good, f"{pct(health_good, total_clients)} good health"),
        ]

        col = 1
        for label, value, subtitle in kpi_data:
            self.add_kpi_card(ws, row, col, label, value, subtitle)
            col += 2

        row += 6

        # --- Health Distribution Pie Chart ---
        chart_row = row
        ws.cell(row=row, column=1, value="Health Distribution")
        ws.cell(row=row, column=1).font = self.styles.get_subtitle_font()
        row += 2

        # Write data table for chart
        health_data = [
            ("Good", summary.get("health_good", 0)),
            ("Fair", summary.get("health_fair", 0)),
            ("Poor", summary.get("health_poor", 0)),
            ("Unknown", summary.get("health_unknown", 0)),
        ]
        data_start_row = row
        for label, count in health_data:
            ws.cell(row=row, column=1, value=label)
            ws.cell(row=row, column=2, value=count)
            row += 1
        data_end_row = row - 1

        # Create pie chart
        if sum(v for _, v in health_data) > 0:
            chart = PieChart()
            chart.title = "Health Distribution"
            labels = Reference(ws, min_col=1, min_row=data_start_row, max_row=data_end_row)
            data = Reference(ws, min_col=2, min_row=data_start_row, max_row=data_end_row)
            chart.add_data(data, titles_from_data=False)
            chart.set_categories(labels)
            chart.dataLabels = DataLabelList()
            chart.dataLabels.showPercent = True
            chart.dataLabels.showVal = True
            chart.width = 12
            chart.height = 8
            ws.add_chart(chart, "D" + str(chart_row + 1))

        row += 2

        # --- Connection Type Pie Chart ---
        chart_row2 = row
        ws.cell(row=row, column=1, value="Connection Type")
        ws.cell(row=row, column=1).font = self.styles.get_subtitle_font()
        row += 2

        type_data = [
            ("Wireless", summary.get("wireless", 0)),
            ("Wired", summary.get("wired", 0)),
        ]
        data_start_row = row
        for label, count in type_data:
            ws.cell(row=row, column=1, value=label)
            ws.cell(row=row, column=2, value=count)
            row += 1
        data_end_row = row - 1

        if sum(v for _, v in type_data) > 0:
            chart2 = PieChart()
            chart2.title = "Connection Type"
            labels2 = Reference(ws, min_col=1, min_row=data_start_row, max_row=data_end_row)
            data2 = Reference(ws, min_col=2, min_row=data_start_row, max_row=data_end_row)
            chart2.add_data(data2, titles_from_data=False)
            chart2.set_categories(labels2)
            chart2.dataLabels = DataLabelList()
            chart2.dataLabels.showPercent = True
            chart2.dataLabels.showVal = True
            chart2.width = 12
            chart2.height = 8
            ws.add_chart(chart2, "D" + str(chart_row2 + 1))

        row += 2

        # --- Top 10 Sites Bar Chart ---
        chart_row3 = row
        ws.cell(row=row, column=1, value="Top 10 Sites by Client Count")
        ws.cell(row=row, column=1).font = self.styles.get_subtitle_font()
        row += 2

        # Calculate site counts
        site_counts: dict[str, int] = {}
        for item in items:
            site = item.get("site_name", "Unknown") or "Unknown"
            site_counts[site] = site_counts.get(site, 0) + 1

        top_sites = sorted(site_counts.items(), key=lambda x: -x[1])[:10]

        # Write data table for chart
        ws.cell(row=row, column=1, value="Site")
        ws.cell(row=row, column=2, value="Clients")
        data_start_row = row + 1

        for site, count in top_sites:
            row += 1
            ws.cell(row=row, column=1, value=site[:30])  # Truncate long names
            ws.cell(row=row, column=2, value=count)
        data_end_row = row

        if top_sites:
            chart3 = BarChart()
            chart3.title = "Top 10 Sites"
            chart3.type = "col"
            chart3.style = 10
            labels3 = Reference(ws, min_col=1, min_row=data_start_row, max_row=data_end_row)
            data3 = Reference(ws, min_col=2, min_row=data_start_row - 1, max_row=data_end_row, titles_from_data=True)
            chart3.add_data(data3, titles_from_data=True)
            chart3.set_categories(labels3)
            chart3.width = 16
            chart3.height = 10
            chart3.shape = 4
            ws.add_chart(chart3, "D" + str(chart_row3 + 1))

        self.auto_fit_columns(ws)

    def _count_sites(self, items: list[dict]) -> int:
        """Count unique sites in items."""
        sites = set()
        for item in items:
            site = item.get("site_name")
            if site:
                sites.add(site)
        return len(sites)  # Return 0 for empty, not 1

    def generate_csv(
        self,
        data: dict[str, Any],
        filters: dict[str, Any] | None = None,
    ) -> str:
        """Generate CSV with client data."""
        items = data.get("items", [])

        rows = []
        for item in items:
            row = {field: item.get(field, "") for field, _ in self.CLIENT_COLUMNS}
            rows.append(row)

        fieldnames = [field for field, _ in self.CLIENT_COLUMNS]
        return self._dict_to_csv(rows, fieldnames)

    def _create_summary_sheet(
        self,
        ws,
        items: list[dict],
        total: int,
        summary: dict[str, Any],
        filters: dict[str, Any] | None,
    ) -> None:
        """Create summary sheet with health KPIs."""
        row = self.add_report_header(
            ws,
            "Network Clients Report",
            f"{total:,} Clients",
            filters,
        )

        # Calculate stats if not provided
        if not summary:
            summary = {
                "total_clients": len(items),
                "connected": sum(1 for c in items if c.get("status") == "Connected"),
                "wired": sum(1 for c in items if c.get("type") == "Wired"),
                "wireless": sum(1 for c in items if c.get("type") == "Wireless"),
                "health_good": sum(1 for c in items if c.get("health") == "Good"),
                "health_fair": sum(1 for c in items if c.get("health") == "Fair"),
                "health_poor": sum(1 for c in items if c.get("health") == "Poor"),
            }

        # KPIs
        ws.cell(row=row, column=1, value="Client Overview")
        ws.cell(row=row, column=1).font = self.styles.get_subtitle_font()
        row += 2

        total_clients = summary.get("total_clients", len(items)) or 1
        connected = summary.get("connected", 0)
        conn_pct = f"{(connected / total_clients * 100):.1f}%"

        kpi_data = [
            ("Total Clients", total_clients, None),
            ("Connected", connected, conn_pct),
            ("Wireless", summary.get("wireless", 0), None),
            ("Wired", summary.get("wired", 0), None),
        ]

        col = 1
        kpi_row = row
        for label, value, subtitle in kpi_data:
            self.add_kpi_card(ws, kpi_row, col, label, value, subtitle)
            col += 2

        row = kpi_row + 6

        # Health breakdown
        ws.cell(row=row, column=1, value="Health Distribution")
        ws.cell(row=row, column=1).font = self.styles.get_subtitle_font()
        row += 2

        health_data = [
            ("Good", summary.get("health_good", 0), "success"),
            ("Fair", summary.get("health_fair", 0), "warning"),
            ("Poor", summary.get("health_poor", 0), "error"),
            ("Unknown", summary.get("health_unknown", 0), "info"),
        ]

        headers = ["Health Status", "Count", "Percentage"]
        self.add_table_headers(ws, headers, row)
        row += 1

        for health, count, status_type in health_data:
            pct = f"{(count / total_clients * 100):.1f}%"
            self.add_data_row(ws, [health, count, pct], row, alternate=(row % 2 == 0))

            if status_type == "success":
                ws.cell(row=row, column=1).fill = self.styles.get_success_fill()
            elif status_type == "warning":
                ws.cell(row=row, column=1).fill = self.styles.get_warning_fill()
            elif status_type == "error":
                ws.cell(row=row, column=1).fill = self.styles.get_error_fill()

            row += 1

        row += 2

        # Status breakdown
        ws.cell(row=row, column=1, value="Connection Status")
        ws.cell(row=row, column=1).font = self.styles.get_subtitle_font()
        row += 2

        status_counts: dict[str, int] = {}
        for item in items:
            status = item.get("status", "Unknown") or "Unknown"
            status_counts[status] = status_counts.get(status, 0) + 1

        headers = ["Status", "Count", "Percentage"]
        self.add_table_headers(ws, headers, row)
        row += 1

        for status, count in sorted(status_counts.items(), key=lambda x: -x[1]):
            pct = f"{(count / total_clients * 100):.1f}%"
            self.add_data_row(ws, [status, count, pct], row, alternate=(row % 2 == 0))
            ws.cell(row=row, column=1).fill = self.styles.get_status_fill(status)
            row += 1

        self.auto_fit_columns(ws)
        self.freeze_panes(ws, row=8)

    def _create_client_list_sheet(self, ws, items: list[dict]) -> None:
        """Create the main client list sheet."""
        row = 1

        headers = [label for _, label in self.CLIENT_COLUMNS]
        self.add_table_headers(ws, headers, row)
        row += 1

        for item in items:
            row_data = [item.get(field, "") for field, _ in self.CLIENT_COLUMNS]
            self.add_data_row(ws, row_data, row, alternate=(row % 2 == 0))

            # Color code health
            health = item.get("health")
            if health:
                ws.cell(row=row, column=4).fill = self.styles.get_status_fill(health)

            # Color code status
            status = item.get("status")
            if status:
                ws.cell(row=row, column=5).fill = self.styles.get_status_fill(status)

            row += 1

        self.auto_fit_columns(ws, max_width=35)
        self.freeze_panes(ws, row=2)

    def _create_site_stats_sheet(self, ws, items: list[dict]) -> None:
        """Create comprehensive per-site statistics sheet with rich details."""
        row = self.add_report_header(
            ws,
            "Site Statistics",
            "Comprehensive Site Analysis",
        )

        # Aggregate detailed stats by site
        site_stats: dict[str, dict] = {}
        for item in items:
            site = item.get("site_name", "Unknown") or "Unknown"
            if site not in site_stats:
                site_stats[site] = {
                    "total": 0,
                    "connected": 0,
                    "disconnected": 0,
                    "wireless": 0,
                    "wired": 0,
                    "good": 0,
                    "fair": 0,
                    "poor": 0,
                    "unknown_health": 0,
                    "networks": set(),
                    "vlans": set(),
                    "auth_methods": set(),
                    "devices": set(),
                    "last_seen": None,
                }
            stats = site_stats[site]
            stats["total"] += 1

            # Connection status
            status = item.get("status")
            if status == "Connected":
                stats["connected"] += 1
            else:
                stats["disconnected"] += 1

            # Connection type
            if item.get("type") == "Wireless":
                stats["wireless"] += 1
            elif item.get("type") == "Wired":
                stats["wired"] += 1

            # Health
            health = (item.get("health") or "").lower()
            if health == "good":
                stats["good"] += 1
            elif health == "fair":
                stats["fair"] += 1
            elif health == "poor":
                stats["poor"] += 1
            else:
                stats["unknown_health"] += 1

            # Networks/SSIDs
            network = item.get("network")
            if network:
                stats["networks"].add(network)

            # VLANs
            vlan = item.get("vlan_id")
            if vlan:
                stats["vlans"].add(str(vlan))

            # Auth methods
            auth = item.get("authentication")
            if auth:
                stats["auth_methods"].add(auth)

            # Connected devices
            connected_to = item.get("connected_to")
            if connected_to:
                stats["devices"].add(connected_to)

            # Track last activity
            last_seen = item.get("last_seen_at")
            if last_seen:
                if stats["last_seen"] is None or last_seen > stats["last_seen"]:
                    stats["last_seen"] = last_seen

        # Calculate health scores
        for site, stats in site_stats.items():
            total = stats["total"] or 1
            # Health score: Good=100, Fair=66, Poor=33, Unknown=0
            known_health = stats["good"] + stats["fair"] + stats["poor"]
            if known_health > 0:
                score = (stats["good"] * 100 + stats["fair"] * 66 + stats["poor"] * 33) / known_health
                stats["health_score"] = round(score, 1)
            else:
                stats["health_score"] = None
            stats["online_rate"] = round(stats["connected"] / total * 100, 1)

        # Full headers with all metrics
        headers = [
            "Site",
            "Total",
            "Connected",
            "Disconnected",
            "Online %",
            "Wireless",
            "Wired",
            "Health Good",
            "Health Fair",
            "Health Poor",
            "Health Score",
            "Networks",
            "VLANs",
            "Auth Methods",
            "Devices",
            "Last Activity",
        ]
        self.add_table_headers(ws, headers, row)
        row += 1

        for site, stats in sorted(site_stats.items(), key=lambda x: -x[1]["total"]):
            self.add_data_row(
                ws,
                [
                    site,
                    stats["total"],
                    stats["connected"],
                    stats["disconnected"],
                    f"{stats['online_rate']}%",
                    stats["wireless"],
                    stats["wired"],
                    stats["good"],
                    stats["fair"],
                    stats["poor"],
                    f"{stats['health_score']}%" if stats["health_score"] is not None else "N/A",
                    len(stats["networks"]),
                    len(stats["vlans"]),
                    len(stats["auth_methods"]),
                    len(stats["devices"]),
                    stats["last_seen"] or "N/A",
                ],
                row,
                alternate=(row % 2 == 0),
            )

            # Color code health score
            health_score = stats["health_score"]
            if health_score is not None:
                if health_score >= 80:
                    ws.cell(row=row, column=11).fill = self.styles.get_success_fill()
                elif health_score >= 50:
                    ws.cell(row=row, column=11).fill = self.styles.get_warning_fill()
                else:
                    ws.cell(row=row, column=11).fill = self.styles.get_error_fill()

            # Highlight low online rates
            if stats["online_rate"] < 50:
                ws.cell(row=row, column=5).fill = self.styles.get_error_fill()
            elif stats["online_rate"] < 80:
                ws.cell(row=row, column=5).fill = self.styles.get_warning_fill()

            # Highlight poor health counts
            if stats["poor"] > 0:
                ws.cell(row=row, column=10).fill = self.styles.get_error_fill()

            row += 1

        row += 2

        # Add site summary bar chart
        ws.cell(row=row, column=1, value="Top Sites by Client Count")
        ws.cell(row=row, column=1).font = self.styles.get_subtitle_font()
        row += 2

        # Write mini-table for chart
        ws.cell(row=row, column=1, value="Site")
        ws.cell(row=row, column=2, value="Clients")
        chart_data_start = row + 1

        top_sites = sorted(site_stats.items(), key=lambda x: -x[1]["total"])[:10]
        for site, stats in top_sites:
            row += 1
            ws.cell(row=row, column=1, value=site[:25])
            ws.cell(row=row, column=2, value=stats["total"])
        chart_data_end = row

        if top_sites:
            chart = BarChart()
            chart.title = "Top 10 Sites"
            chart.type = "col"
            chart.style = 10
            labels = Reference(ws, min_col=1, min_row=chart_data_start, max_row=chart_data_end)
            data = Reference(ws, min_col=2, min_row=chart_data_start - 1, max_row=chart_data_end, titles_from_data=True)
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(labels)
            chart.width = 14
            chart.height = 8
            ws.add_chart(chart, "D" + str(chart_data_start - 2))

        self.auto_fit_columns(ws)
        self.freeze_panes(ws, row=6)

    def _create_network_analysis_sheet(self, ws, items: list[dict]) -> None:
        """Create network segmentation analysis sheet."""
        row = self.add_report_header(
            ws,
            "Network Analysis",
            "VLAN and Network Distribution",
        )

        # VLAN breakdown
        ws.cell(row=row, column=1, value="VLAN Distribution")
        ws.cell(row=row, column=1).font = self.styles.get_subtitle_font()
        row += 2

        vlan_counts: dict[str, int] = {}
        for item in items:
            vlan = item.get("vlan_id", "Unknown") or "Unknown"
            vlan_counts[str(vlan)] = vlan_counts.get(str(vlan), 0) + 1

        headers = ["VLAN", "Client Count", "Percentage"]
        self.add_table_headers(ws, headers, row)
        row += 1

        total_items = max(len(items), 1)
        for vlan, count in sorted(vlan_counts.items(), key=lambda x: -x[1])[:20]:
            pct = f"{(count / total_items * 100):.1f}%"
            self.add_data_row(ws, [vlan, count, pct], row, alternate=(row % 2 == 0))
            row += 1

        if len(vlan_counts) > 20:
            ws.cell(row=row, column=1, value=f"... and {len(vlan_counts) - 20} more VLANs")
            ws.cell(row=row, column=1).font = Font(italic=True)
            row += 1

        row += 2

        # Network/SSID breakdown (for wireless)
        ws.cell(row=row, column=1, value="Wireless Network (SSID) Distribution")
        ws.cell(row=row, column=1).font = self.styles.get_subtitle_font()
        row += 2

        network_counts: dict[str, int] = {}
        for item in items:
            if item.get("type") == "Wireless":
                network = item.get("network", "Unknown") or "Unknown"
                network_counts[network] = network_counts.get(network, 0) + 1

        if network_counts:
            headers = ["Network/SSID", "Client Count", "Percentage"]
            self.add_table_headers(ws, headers, row)
            row += 1

            wireless_total = max(sum(network_counts.values()), 1)
            for network, count in sorted(network_counts.items(), key=lambda x: -x[1])[:15]:
                pct = f"{(count / wireless_total * 100):.1f}%"
                self.add_data_row(ws, [network, count, pct], row, alternate=(row % 2 == 0))
                row += 1
        else:
            ws.cell(row=row, column=1, value="No wireless clients found.")
            ws.cell(row=row, column=1).font = Font(italic=True)

        self.auto_fit_columns(ws)
        self.freeze_panes(ws, row=6)
