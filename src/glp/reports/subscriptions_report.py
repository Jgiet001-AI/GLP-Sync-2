"""Subscription report generator.

Generates comprehensive subscription reports with:
- License utilization analysis
- Capacity planning data
- Renewal timeline
- Device assignment coverage
"""

import json
import logging
from typing import Any

from openpyxl.styles import Font

from .generator import BaseReportGenerator

logger = logging.getLogger(__name__)


class SubscriptionsReportGenerator(BaseReportGenerator):
    """Generate detailed subscription and license reports."""

    SUBSCRIPTION_COLUMNS = [
        ("key", "Subscription Key"),
        ("subscription_type", "Type"),
        ("subscription_status", "Status"),
        ("tier", "Tier"),
        ("sku", "SKU"),
        ("quantity", "Total Licenses"),
        ("used_quantity", "Used"),
        ("available_quantity", "Available"),
        ("device_count", "Devices"),
        ("start_time", "Start Date"),
        ("end_time", "End Date"),
        ("days_remaining", "Days Remaining"),
        ("is_eval", "Evaluation"),
    ]

    def generate_excel(
        self,
        data: dict[str, Any],
        filters: dict[str, Any] | None = None,
    ) -> bytes:
        """Generate multi-sheet subscription Excel report.

        Sheets:
        1. Summary - Utilization overview
        2. Subscription List - Full inventory
        3. Capacity Analysis - License utilization by type
        4. Renewal Planning - Expiration timeline
        """
        wb = self.create_workbook()

        items = data.get("items", [])
        total = data.get("total", len(items))

        # Sheet 1: Summary
        ws_summary = wb.active
        ws_summary.title = "Summary"
        self._create_summary_sheet(ws_summary, items, total, filters)

        # Sheet 2: Subscription List
        ws_list = wb.create_sheet("Subscription List")
        self._create_subscription_list_sheet(ws_list, items)

        # Sheet 3: Capacity Analysis
        ws_capacity = wb.create_sheet("Capacity Analysis")
        self._create_capacity_sheet(ws_capacity, items)

        # Sheet 4: Renewal Planning
        ws_renewal = wb.create_sheet("Renewal Planning")
        self._create_renewal_sheet(ws_renewal, items)

        return self._workbook_to_bytes(wb)

    def generate_csv(
        self,
        data: dict[str, Any],
        filters: dict[str, Any] | None = None,
    ) -> str:
        """Generate CSV with subscription data."""
        items = data.get("items", [])

        rows = []
        for item in items:
            row = {}
            for field, _ in self.SUBSCRIPTION_COLUMNS:
                value = item.get(field, "")
                if field == "subscription_type" and value:
                    value = value.replace("CENTRAL_", "")
                if field == "is_eval":
                    value = "Yes" if value else "No"
                row[field] = value
            rows.append(row)

        fieldnames = [field for field, _ in self.SUBSCRIPTION_COLUMNS]
        return self._dict_to_csv(rows, fieldnames)

    def generate_json(
        self,
        data: dict[str, Any],
        filters: dict[str, Any] | None = None,
    ) -> str:
        """Generate JSON with subscription data."""
        items = data.get("items", [])

        # Return JSON array of subscriptions
        return json.dumps(items, indent=2, default=str)

    def _create_summary_sheet(
        self,
        ws,
        items: list[dict],
        total: int,
        filters: dict[str, Any] | None,
    ) -> None:
        """Create summary sheet with utilization KPIs."""
        row = self.add_report_header(
            ws,
            "Subscription Report",
            f"{total:,} Subscriptions",
            filters,
        )

        # Calculate totals
        total_licenses = sum(s.get("quantity", 0) for s in items)
        used_licenses = sum(s.get("used_quantity", 0) for s in items)
        utilization = (used_licenses / max(total_licenses, 1)) * 100

        active = sum(1 for s in items if s.get("subscription_status") == "STARTED")
        expired = sum(1 for s in items if s.get("subscription_status") in ("ENDED", "CANCELLED"))
        expiring_soon = sum(1 for s in items if 0 < (s.get("days_remaining") or 999) <= 90)

        # KPIs
        ws.cell(row=row, column=1, value="License Utilization Overview")
        ws.cell(row=row, column=1).font = self.styles.get_subtitle_font()
        row += 2

        kpi_data = [
            ("Total Licenses", f"{total_licenses:,}", None),
            ("Utilization", f"{utilization:.1f}%", f"{used_licenses:,} used"),
            ("Active Subscriptions", active, f"{expired} expired"),
            ("Expiring Soon", expiring_soon, "Within 90 days"),
        ]

        col = 1
        kpi_row = row
        for label, value, subtitle in kpi_data:
            self.add_kpi_card(ws, kpi_row, col, label, value, subtitle)
            col += 2

        row = kpi_row + 6

        # Status breakdown
        ws.cell(row=row, column=1, value="Status Breakdown")
        ws.cell(row=row, column=1).font = self.styles.get_subtitle_font()
        row += 2

        status_counts: dict[str, int] = {}
        for item in items:
            status = item.get("subscription_status", "Unknown") or "Unknown"
            status_counts[status] = status_counts.get(status, 0) + 1

        headers = ["Status", "Count", "Percentage"]
        self.add_table_headers(ws, headers, row)
        row += 1

        for status, count in sorted(status_counts.items(), key=lambda x: -x[1]):
            pct = f"{(count / max(len(items), 1) * 100):.1f}%"
            self.add_data_row(ws, [status, count, pct], row, alternate=(row % 2 == 0))
            ws.cell(row=row, column=1).fill = self.styles.get_status_fill(status)
            row += 1

        self.auto_fit_columns(ws)
        self.freeze_panes(ws, row=8)

    def _create_subscription_list_sheet(self, ws, items: list[dict]) -> None:
        """Create the main subscription list sheet."""
        row = 1

        headers = [label for _, label in self.SUBSCRIPTION_COLUMNS]
        self.add_table_headers(ws, headers, row)
        row += 1

        for item in items:
            row_data = []
            for field, _ in self.SUBSCRIPTION_COLUMNS:
                value = item.get(field, "")

                if field == "subscription_type" and value:
                    value = value.replace("CENTRAL_", "")
                elif field == "is_eval":
                    value = "Yes" if value else "No"

                row_data.append(value)

            self.add_data_row(ws, row_data, row, alternate=(row % 2 == 0))

            # Color code status
            status = item.get("subscription_status")
            if status:
                ws.cell(row=row, column=3).fill = self.styles.get_status_fill(status)

            # Color code days remaining
            days = item.get("days_remaining")
            if days is not None:
                ws.cell(row=row, column=12).fill = self.styles.get_urgency_fill(days)

            # Highlight high utilization
            qty = item.get("quantity", 0) or 1
            used = item.get("used_quantity", 0)
            if (used / qty * 100) >= 90:
                ws.cell(row=row, column=7).fill = self.styles.get_warning_fill()

            row += 1

        self.auto_fit_columns(ws, max_width=35)
        self.freeze_panes(ws, row=2)

    def _create_capacity_sheet(self, ws, items: list[dict]) -> None:
        """Create capacity analysis sheet by subscription type."""
        row = self.add_report_header(
            ws,
            "Capacity Analysis",
            "License Utilization by Type",
        )

        # Aggregate by type
        type_stats: dict[str, dict] = {}
        for item in items:
            sub_type = (item.get("subscription_type") or "Unknown").replace("CENTRAL_", "")
            if sub_type not in type_stats:
                type_stats[sub_type] = {
                    "count": 0,
                    "total": 0,
                    "used": 0,
                    "available": 0,
                    "devices": 0,
                }
            type_stats[sub_type]["count"] += 1
            type_stats[sub_type]["total"] += item.get("quantity", 0)
            type_stats[sub_type]["used"] += item.get("used_quantity", 0)
            type_stats[sub_type]["available"] += item.get("available_quantity", 0)
            type_stats[sub_type]["devices"] += item.get("device_count", 0)

        headers = [
            "Subscription Type",
            "Subscriptions",
            "Total Licenses",
            "Used",
            "Available",
            "Utilization",
            "Devices",
        ]
        self.add_table_headers(ws, headers, row)
        row += 1

        for sub_type, stats in sorted(type_stats.items(), key=lambda x: -x[1]["total"]):
            total = stats["total"] or 1
            utilization = f"{(stats['used'] / total * 100):.1f}%"

            self.add_data_row(
                ws,
                [
                    sub_type,
                    stats["count"],
                    stats["total"],
                    stats["used"],
                    stats["available"],
                    utilization,
                    stats["devices"],
                ],
                row,
                alternate=(row % 2 == 0),
            )

            # Highlight utilization
            util_pct = stats["used"] / total * 100
            if util_pct >= 90:
                ws.cell(row=row, column=6).fill = self.styles.get_warning_fill()
            elif util_pct < 50:
                ws.cell(row=row, column=6).fill = self.styles.get_info_fill()

            row += 1

        # Grand total
        row += 1
        grand_total = sum(s["total"] for s in type_stats.values())
        grand_used = sum(s["used"] for s in type_stats.values())
        grand_available = sum(s["available"] for s in type_stats.values())
        grand_devices = sum(s["devices"] for s in type_stats.values())
        grand_util = f"{(grand_used / max(grand_total, 1) * 100):.1f}%"

        self.add_data_row(
            ws,
            ["TOTAL", len(items), grand_total, grand_used, grand_available, grand_util, grand_devices],
            row,
        )
        for col in range(1, 8):
            ws.cell(row=row, column=col).font = Font(bold=True)
            ws.cell(row=row, column=col).fill = self.styles.get_hpe_green_light_fill()

        self.auto_fit_columns(ws)
        self.freeze_panes(ws, row=6)

    def _create_renewal_sheet(self, ws, items: list[dict]) -> None:
        """Create renewal planning sheet sorted by expiration."""
        row = self.add_report_header(
            ws,
            "Renewal Planning",
            "Subscriptions Sorted by Expiration Date",
        )

        # Filter and sort active subscriptions
        active_subs = [
            s for s in items
            if s.get("subscription_status") == "STARTED" and s.get("days_remaining") is not None
        ]
        sorted_subs = sorted(active_subs, key=lambda x: x.get("days_remaining", 999))

        if not sorted_subs:
            ws.cell(row=row, column=1, value="No active subscriptions with expiration dates.")
            ws.cell(row=row, column=1).font = Font(italic=True)
            return

        headers = [
            "Subscription Key",
            "Type",
            "Tier",
            "Licenses",
            "End Date",
            "Days Remaining",
            "Urgency",
            "Devices Affected",
        ]
        self.add_table_headers(ws, headers, row)
        row += 1

        for item in sorted_subs:
            days = item.get("days_remaining", 0)

            if days <= 7:
                urgency = "CRITICAL"
            elif days <= 30:
                urgency = "HIGH"
            elif days <= 90:
                urgency = "MEDIUM"
            else:
                urgency = "LOW"

            self.add_data_row(
                ws,
                [
                    item.get("key"),
                    (item.get("subscription_type") or "").replace("CENTRAL_", ""),
                    item.get("tier"),
                    item.get("quantity"),
                    item.get("end_time"),
                    days,
                    urgency,
                    item.get("device_count"),
                ],
                row,
                alternate=(row % 2 == 0),
            )

            # Color code entire row by urgency
            urgency_fill = self.styles.get_urgency_fill(days)
            for col in range(1, 9):
                ws.cell(row=row, column=col).fill = urgency_fill

            row += 1

        self.auto_fit_columns(ws, max_width=35)
        self.freeze_panes(ws, row=6)
