"""Base report generator with common functionality.

This module provides the foundation for all report generators with:
- Excel formula injection protection
- Thread-safe async generation
- Common styling and formatting
- CSV and Excel output support
"""

import csv
import io
import json
import logging
import re
from abc import ABC, abstractmethod
from datetime import datetime
from typing import Any

import anyio
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from .styles import ExcelStyles

logger = logging.getLogger(__name__)


class BaseReportGenerator(ABC):
    """Abstract base class for report generators.

    Provides common functionality for:
    - Excel formula injection sanitization
    - Column width auto-fitting
    - Header row creation
    - Async generation wrapper
    """

    # Characters that could trigger Excel formula interpretation
    FORMULA_CHARS = ("=", "+", "-", "@", "\t", "\r", "\n")

    def __init__(self) -> None:
        self.styles = ExcelStyles

    @classmethod
    def sanitize_cell_value(cls, value: Any) -> Any:
        """Sanitize cell value to prevent Excel formula injection.

        Args:
            value: The value to sanitize

        Returns:
            Sanitized value safe for Excel cells
        """
        if value is None:
            return ""

        if isinstance(value, str):
            # Check if starts with formula-triggering characters
            if value and value[0] in cls.FORMULA_CHARS:
                # Prefix with apostrophe to force text interpretation
                return f"'{value}"
            # Also sanitize embedded formulas
            if "=" in value and re.match(r".*=\s*[A-Za-z]+\(", value):
                return f"'{value}"
            return value

        # Return non-string values as-is
        return value

    @classmethod
    def sanitize_dict(cls, data: dict[str, Any]) -> dict[str, Any]:
        """Sanitize all string values in a dictionary."""
        return {k: cls.sanitize_cell_value(v) for k, v in data.items()}

    def create_workbook(self) -> Workbook:
        """Create a new workbook with named styles registered."""
        wb = Workbook()
        # Register named styles
        for style in self.styles.create_named_styles():
            try:
                wb.add_named_style(style)
            except ValueError:
                # Style already exists
                pass
        return wb

    def add_report_header(
        self,
        ws: Worksheet,
        title: str,
        subtitle: str | None = None,
        filters: dict[str, Any] | None = None,
    ) -> int:
        """Add a report header section to a worksheet.

        Args:
            ws: The worksheet to add the header to
            title: Main report title
            subtitle: Optional subtitle
            filters: Optional filters applied to the report

        Returns:
            The next available row number
        """
        row = 1

        # Title
        ws.cell(row=row, column=1, value=title)
        ws.cell(row=row, column=1).font = self.styles.get_title_font()
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        row += 1

        # Subtitle
        if subtitle:
            ws.cell(row=row, column=1, value=subtitle)
            ws.cell(row=row, column=1).font = self.styles.get_subtitle_font()
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
            row += 1

        row += 1  # Blank row

        # Report metadata
        ws.cell(row=row, column=1, value="Generated:")
        ws.cell(row=row, column=2, value=datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
        row += 1

        # Applied filters
        if filters:
            active_filters = {k: v for k, v in filters.items() if v is not None}
            if active_filters:
                ws.cell(row=row, column=1, value="Filters Applied:")
                filter_str = ", ".join(f"{k}={v}" for k, v in active_filters.items())
                ws.cell(row=row, column=2, value=filter_str)
                row += 1

        row += 1  # Blank row before data
        return row

    def add_table_headers(
        self,
        ws: Worksheet,
        headers: list[str],
        start_row: int,
        start_col: int = 1,
    ) -> None:
        """Add styled table headers to a worksheet.

        Args:
            ws: The worksheet
            headers: List of header strings
            start_row: Starting row number
            start_col: Starting column number
        """
        header_fill = self.styles.get_header_fill()
        header_font = self.styles.get_header_font()
        alignment = self.styles.get_center_alignment()
        border = self.styles.THIN_BORDER

        for col, header in enumerate(headers, start=start_col):
            cell = ws.cell(row=start_row, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = alignment
            cell.border = border

    def add_data_row(
        self,
        ws: Worksheet,
        row_data: list[Any],
        row_num: int,
        start_col: int = 1,
        alternate: bool = False,
    ) -> None:
        """Add a data row to a worksheet with styling.

        Args:
            ws: The worksheet
            row_data: List of values for the row
            row_num: Row number to write to
            start_col: Starting column number
            alternate: Whether to use alternate row styling
        """
        border = self.styles.THIN_BORDER
        alignment = Alignment(vertical="center", wrap_text=True)
        fill = self.styles.get_alternate_row_fill() if alternate else None

        for col, value in enumerate(row_data, start=start_col):
            cell = ws.cell(row=row_num, column=col, value=self.sanitize_cell_value(value))
            cell.border = border
            cell.alignment = alignment
            if fill:
                cell.fill = fill

    def auto_fit_columns(self, ws: Worksheet, min_width: int = 10, max_width: int = 50) -> None:
        """Auto-fit column widths based on content.

        Args:
            ws: The worksheet
            min_width: Minimum column width
            max_width: Maximum column width
        """
        for column_cells in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column_cells[0].column)

            for cell in column_cells:
                try:
                    if cell.value:
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = cell_length
                except (TypeError, AttributeError):
                    pass

            adjusted_width = min(max(max_length + 2, min_width), max_width)
            ws.column_dimensions[column_letter].width = adjusted_width

    def freeze_panes(self, ws: Worksheet, row: int = 2, column: int = 1) -> None:
        """Freeze panes at the specified position.

        Args:
            ws: The worksheet
            row: Row to freeze at (rows above will be frozen)
            column: Column to freeze at (columns to the left will be frozen)
        """
        ws.freeze_panes = ws.cell(row=row, column=column)

    def add_kpi_card(
        self,
        ws: Worksheet,
        row: int,
        col: int,
        label: str,
        value: Any,
        subtitle: str | None = None,
    ) -> int:
        """Add a KPI card to the worksheet.

        Args:
            ws: The worksheet
            row: Starting row
            col: Starting column
            label: KPI label
            value: KPI value
            subtitle: Optional subtitle/context

        Returns:
            Next available row
        """
        # Label
        ws.cell(row=row, column=col, value=label)
        ws.cell(row=row, column=col).font = self.styles.get_kpi_label_font()

        # Value
        row += 1
        ws.cell(row=row, column=col, value=value)
        ws.cell(row=row, column=col).font = self.styles.get_kpi_value_font()

        # Subtitle
        if subtitle:
            row += 1
            ws.cell(row=row, column=col, value=subtitle)
            ws.cell(row=row, column=col).font = Font(size=9, color=self.styles.DARK_GRAY)

        return row + 2

    @abstractmethod
    def generate_excel(self, data: dict[str, Any], filters: dict[str, Any] | None = None) -> bytes:
        """Generate Excel report bytes.

        Args:
            data: Report data
            filters: Optional filters applied

        Returns:
            Excel file as bytes
        """
        pass

    @abstractmethod
    def generate_csv(self, data: dict[str, Any], filters: dict[str, Any] | None = None) -> str:
        """Generate CSV report string.

        Args:
            data: Report data
            filters: Optional filters applied

        Returns:
            CSV content as string
        """
        pass

    @abstractmethod
    def generate_json(self, data: dict[str, Any], filters: dict[str, Any] | None = None) -> str:
        """Generate JSON report string.

        Args:
            data: Report data
            filters: Optional filters applied

        Returns:
            JSON content as string
        """
        pass

    async def generate_excel_async(
        self,
        data: dict[str, Any],
        filters: dict[str, Any] | None = None,
    ) -> bytes:
        """Generate Excel report asynchronously (non-blocking).

        Uses anyio.to_thread to offload heavy workbook generation
        from the async event loop.

        Args:
            data: Report data
            filters: Optional filters applied

        Returns:
            Excel file as bytes
        """
        return await anyio.to_thread.run_sync(
            lambda: self.generate_excel(data, filters)
        )

    async def generate_csv_async(
        self,
        data: dict[str, Any],
        filters: dict[str, Any] | None = None,
    ) -> str:
        """Generate CSV report asynchronously (non-blocking).

        Args:
            data: Report data
            filters: Optional filters applied

        Returns:
            CSV content as string
        """
        return await anyio.to_thread.run_sync(
            lambda: self.generate_csv(data, filters)
        )

    async def generate_json_async(
        self,
        data: dict[str, Any],
        filters: dict[str, Any] | None = None,
    ) -> str:
        """Generate JSON report asynchronously (non-blocking).

        Args:
            data: Report data
            filters: Optional filters applied

        Returns:
            JSON content as string
        """
        return await anyio.to_thread.run_sync(
            lambda: self.generate_json(data, filters)
        )

    def _workbook_to_bytes(self, wb: Workbook) -> bytes:
        """Convert workbook to bytes."""
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()

    def _dict_to_csv(
        self,
        data: list[dict[str, Any]],
        fieldnames: list[str] | None = None,
    ) -> str:
        """Convert list of dicts to CSV string.

        Args:
            data: List of dictionaries
            fieldnames: Optional list of field names (order and selection)

        Returns:
            CSV string
        """
        if not data:
            return ""

        if fieldnames is None:
            fieldnames = list(data[0].keys())

        output = io.StringIO()
        writer = csv.DictWriter(
            output,
            fieldnames=fieldnames,
            extrasaction="ignore",
            quoting=csv.QUOTE_MINIMAL,
        )
        writer.writeheader()

        for row in data:
            # Sanitize values for CSV
            sanitized_row = {
                k: self.sanitize_cell_value(v) for k, v in row.items()
            }
            writer.writerow(sanitized_row)

        return output.getvalue()
