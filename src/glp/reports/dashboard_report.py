"""Dashboard report generator for executive summaries.

Generates comprehensive Excel reports with:
- Executive summary with KPIs
- Device inventory breakdown
- Subscription analysis
- Expiring items list
- Sync history
"""

import logging
from typing import Any

from openpyxl.styles import Font

from .generator import BaseReportGenerator

logger = logging.getLogger(__name__)


class DashboardReportGenerator(BaseReportGenerator):
    """Generate beautiful dashboard executive reports."""

    def generate_excel(
        self,
        data: dict[str, Any],
        filters: dict[str, Any] | None = None,
    ) -> bytes:
        """Generate a multi-sheet Excel dashboard report.

        Sheets:
        1. Executive Summary - KPIs and key insights
        2. Device Inventory - Breakdown by type and region
        3. Subscription Analysis - Utilization and capacity
        4. Expiring Items - Items expiring soon
        5. Sync History - Recent sync operations
        """
        wb = self.create_workbook()

        # Sheet 1: Executive Summary
        self._create_executive_summary(wb.active, data)
        wb.active.title = "Executive Summary"

        # Sheet 2: Device Inventory
        ws_devices = wb.create_sheet("Device Inventory")
        self._create_device_inventory_sheet(ws_devices, data)

        # Sheet 3: Subscription Analysis
        ws_subs = wb.create_sheet("Subscription Analysis")
        self._create_subscription_sheet(ws_subs, data)

        # Sheet 4: Expiring Items
        ws_expiring = wb.create_sheet("Expiring Items")
        self._create_expiring_items_sheet(ws_expiring, data)

        # Sheet 5: Sync History
        ws_sync = wb.create_sheet("Sync History")
        self._create_sync_history_sheet(ws_sync, data)

        return self._workbook_to_bytes(wb)

    def generate_csv(
        self,
        data: dict[str, Any],
        filters: dict[str, Any] | None = None,
    ) -> str:
        """Generate CSV with dashboard summary data."""
        rows = []

        # Device stats
        device_stats = data.get("device_stats", {})
        rows.append({
            "Category": "Devices",
            "Metric": "Total",
            "Value": device_stats.get("total", 0),
        })
        rows.append({
            "Category": "Devices",
            "Metric": "Assigned",
            "Value": device_stats.get("assigned", 0),
        })
        rows.append({
            "Category": "Devices",
            "Metric": "Unassigned",
            "Value": device_stats.get("unassigned", 0),
        })

        # Subscription stats
        sub_stats = data.get("subscription_stats", {})
        rows.append({
            "Category": "Subscriptions",
            "Metric": "Active",
            "Value": sub_stats.get("active", 0),
        })
        rows.append({
            "Category": "Subscriptions",
            "Metric": "Total Licenses",
            "Value": sub_stats.get("total_licenses", 0),
        })
        rows.append({
            "Category": "Subscriptions",
            "Metric": "Utilization %",
            "Value": sub_stats.get("utilization_percent", 0),
        })
        rows.append({
            "Category": "Subscriptions",
            "Metric": "Expiring Soon",
            "Value": sub_stats.get("expiring_soon", 0),
        })

        return self._dict_to_csv(rows, ["Category", "Metric", "Value"])

    def _create_executive_summary(self, ws, data: dict[str, Any]) -> None:
        """Create the executive summary sheet with KPIs."""
        row = self.add_report_header(
            ws,
            "HPE GreenLake Dashboard Report",
            "Executive Summary & Inventory Overview",
        )

        # KPI Section
        ws.cell(row=row, column=1, value="Key Performance Indicators")
        ws.cell(row=row, column=1).font = self.styles.get_subtitle_font()
        row += 2

        device_stats = data.get("device_stats", {})
        sub_stats = data.get("subscription_stats", {})

        # KPI Cards - Row 1
        kpi_data = [
            ("Total Devices", device_stats.get("total", 0), f"{device_stats.get('assigned', 0)} assigned"),
            ("Active Subscriptions", sub_stats.get("active", 0), f"{sub_stats.get('total_licenses', 0):,} total licenses"),
            ("License Utilization", f"{sub_stats.get('utilization_percent', 0)}%", f"{sub_stats.get('total_licenses', 0) - sub_stats.get('available_licenses', 0):,} used"),
            ("Expiring Soon", sub_stats.get("expiring_soon", 0), "Within 90 days"),
        ]

        col = 1
        kpi_row = row
        for label, value, subtitle in kpi_data:
            self.add_kpi_card(ws, kpi_row, col, label, value, subtitle)
            col += 2

        row = kpi_row + 6

        # Device Status Summary
        ws.cell(row=row, column=1, value="Device Status Summary")
        ws.cell(row=row, column=1).font = self.styles.get_subtitle_font()
        row += 2

        status_headers = ["Status", "Count", "Percentage"]
        self.add_table_headers(ws, status_headers, row)
        row += 1

        total_devices = device_stats.get("total", 1) or 1
        status_data = [
            ("Assigned", device_stats.get("assigned", 0)),
            ("Unassigned", device_stats.get("unassigned", 0)),
            ("Archived", device_stats.get("archived", 0)),
        ]

        for status, count in status_data:
            pct = f"{(count / total_devices * 100):.1f}%"
            self.add_data_row(ws, [status, count, pct], row, alternate=(row % 2 == 0))

            # Color code based on status
            if status == "Assigned":
                ws.cell(row=row, column=1).fill = self.styles.get_success_fill()
            elif status == "Unassigned":
                ws.cell(row=row, column=1).fill = self.styles.get_warning_fill()
            row += 1

        self.auto_fit_columns(ws)
        self.freeze_panes(ws, row=8)

    def _create_device_inventory_sheet(self, ws, data: dict[str, Any]) -> None:
        """Create device inventory breakdown sheet."""
        row = self.add_report_header(ws, "Device Inventory", "Breakdown by Type and Region")

        # Device by Type
        ws.cell(row=row, column=1, value="Devices by Type")
        ws.cell(row=row, column=1).font = self.styles.get_subtitle_font()
        row += 2

        type_headers = ["Device Type", "Total", "Assigned", "Unassigned", "Assignment Rate"]
        self.add_table_headers(ws, type_headers, row)
        row += 1

        device_by_type = data.get("device_by_type", [])
        for item in device_by_type:
            total = item.get("count", 0) or 1
            assigned = item.get("assigned", 0)
            unassigned = item.get("unassigned", 0)
            rate = f"{(assigned / total * 100):.1f}%"

            self.add_data_row(
                ws,
                [item.get("device_type"), total, assigned, unassigned, rate],
                row,
                alternate=(row % 2 == 0),
            )
            row += 1

        row += 2

        # Device by Region
        ws.cell(row=row, column=1, value="Devices by Region")
        ws.cell(row=row, column=1).font = self.styles.get_subtitle_font()
        row += 2

        region_headers = ["Region", "Device Count", "Share"]
        self.add_table_headers(ws, region_headers, row)
        row += 1

        device_by_region = data.get("device_by_region", [])
        total_all = sum(item.get("count", 0) for item in device_by_region) or 1

        for item in device_by_region:
            count = item.get("count", 0)
            share = f"{(count / total_all * 100):.1f}%"
            self.add_data_row(
                ws,
                [item.get("region"), count, share],
                row,
                alternate=(row % 2 == 0),
            )
            row += 1

        self.auto_fit_columns(ws)
        self.freeze_panes(ws, row=6)

    def _create_subscription_sheet(self, ws, data: dict[str, Any]) -> None:
        """Create subscription analysis sheet."""
        row = self.add_report_header(ws, "Subscription Analysis", "License Utilization & Capacity")

        # Overall stats
        sub_stats = data.get("subscription_stats", {})

        ws.cell(row=row, column=1, value="License Capacity Overview")
        ws.cell(row=row, column=1).font = self.styles.get_subtitle_font()
        row += 2

        capacity_headers = ["Metric", "Value"]
        self.add_table_headers(ws, capacity_headers, row)
        row += 1

        total_licenses = sub_stats.get("total_licenses", 0)
        available = sub_stats.get("available_licenses", 0)
        used = total_licenses - available
        utilization = sub_stats.get("utilization_percent", 0)

        capacity_data = [
            ("Total Licenses", f"{total_licenses:,}"),
            ("Used Licenses", f"{used:,}"),
            ("Available Licenses", f"{available:,}"),
            ("Utilization Rate", f"{utilization}%"),
            ("Active Subscriptions", sub_stats.get("active", 0)),
            ("Expiring Soon (90 days)", sub_stats.get("expiring_soon", 0)),
        ]

        for metric, value in capacity_data:
            self.add_data_row(ws, [metric, value], row, alternate=(row % 2 == 0))
            row += 1

        row += 2

        # Subscription by Type
        ws.cell(row=row, column=1, value="Subscriptions by Type")
        ws.cell(row=row, column=1).font = self.styles.get_subtitle_font()
        row += 2

        type_headers = ["Type", "Count", "Total Licenses", "Available", "Utilization"]
        self.add_table_headers(ws, type_headers, row)
        row += 1

        sub_by_type = data.get("subscription_by_type", [])
        for item in sub_by_type:
            total_qty = item.get("total_quantity", 0) or 1
            avail_qty = item.get("available_quantity", 0)
            used_qty = total_qty - avail_qty
            util_rate = f"{(used_qty / total_qty * 100):.1f}%"

            self.add_data_row(
                ws,
                [
                    item.get("subscription_type", "").replace("CENTRAL_", ""),
                    item.get("count", 0),
                    total_qty,
                    avail_qty,
                    util_rate,
                ],
                row,
                alternate=(row % 2 == 0),
            )

            # Highlight high utilization
            if (used_qty / total_qty * 100) >= 90:
                ws.cell(row=row, column=5).fill = self.styles.get_warning_fill()

            row += 1

        self.auto_fit_columns(ws)
        self.freeze_panes(ws, row=6)

    def _create_expiring_items_sheet(self, ws, data: dict[str, Any]) -> None:
        """Create expiring items sheet with urgency highlighting."""
        row = self.add_report_header(
            ws,
            "Expiring Items",
            "Devices and Subscriptions Expiring Within 90 Days",
        )

        expiring_items = data.get("expiring_items", [])

        if not expiring_items:
            ws.cell(row=row, column=1, value="No items expiring within 90 days.")
            ws.cell(row=row, column=1).font = Font(size=12, italic=True, color=self.styles.HPE_GREEN)
            return

        headers = ["Type", "Identifier", "Category", "Expiration Date", "Days Remaining", "Urgency"]
        self.add_table_headers(ws, headers, row)
        row += 1

        # Sort by days remaining
        sorted_items = sorted(expiring_items, key=lambda x: x.get("days_remaining", 999))

        for item in sorted_items:
            days = item.get("days_remaining", 0)

            if days <= 7:
                urgency = "CRITICAL"
            elif days <= 30:
                urgency = "HIGH"
            elif days <= 90:
                urgency = "MEDIUM"
            else:
                urgency = "LOW"

            row_data = [
                "Device" if item.get("item_type") == "device" else "Subscription",
                item.get("identifier", ""),
                item.get("sub_type", ""),
                item.get("end_time", ""),
                days,
                urgency,
            ]
            self.add_data_row(ws, row_data, row, alternate=(row % 2 == 0))

            # Apply urgency fill to the entire row
            urgency_fill = self.styles.get_urgency_fill(days)
            for col in range(1, 7):
                ws.cell(row=row, column=col).fill = urgency_fill

            row += 1

        self.auto_fit_columns(ws)
        self.freeze_panes(ws, row=6)

    def _create_sync_history_sheet(self, ws, data: dict[str, Any]) -> None:
        """Create sync history sheet."""
        row = self.add_report_header(ws, "Sync History", "Recent Synchronization Operations")

        sync_history = data.get("sync_history", [])

        if not sync_history:
            ws.cell(row=row, column=1, value="No sync history available.")
            ws.cell(row=row, column=1).font = Font(size=12, italic=True)
            return

        headers = [
            "Resource Type",
            "Started At",
            "Status",
            "Records Fetched",
            "Inserted",
            "Updated",
            "Errors",
            "Duration (ms)",
        ]
        self.add_table_headers(ws, headers, row)
        row += 1

        for item in sync_history:
            status = item.get("status", "")
            row_data = [
                item.get("resource_type", ""),
                item.get("started_at", ""),
                status,
                item.get("records_fetched", 0),
                item.get("records_inserted", 0),
                item.get("records_updated", 0),
                item.get("records_errors", 0),
                item.get("duration_ms", 0),
            ]
            self.add_data_row(ws, row_data, row, alternate=(row % 2 == 0))

            # Color code status
            status_cell = ws.cell(row=row, column=3)
            status_cell.fill = self.styles.get_status_fill(status)

            row += 1

        self.auto_fit_columns(ws)
        self.freeze_panes(ws, row=6)
