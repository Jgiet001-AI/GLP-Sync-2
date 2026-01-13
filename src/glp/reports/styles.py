"""Excel styling definitions for beautiful reports.

This module provides consistent HPE-branded styling for Excel reports
with professional formatting and conditional formatting rules.
"""

from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    NamedStyle,
    PatternFill,
    Side,
)


class ExcelStyles:
    """Excel styling constants and factory methods for HPE-branded reports."""

    # HPE Brand Colors
    HPE_GREEN = "01A982"
    HPE_GREEN_LIGHT = "E6F9F3"
    HPE_PURPLE = "7630EA"
    HPE_PURPLE_LIGHT = "F3EDFD"

    # Status Colors
    SUCCESS_GREEN = "C6EFCE"
    SUCCESS_GREEN_DARK = "006100"
    WARNING_AMBER = "FFEB9C"
    WARNING_AMBER_DARK = "9C5700"
    ERROR_RED = "FFC7CE"
    ERROR_RED_DARK = "9C0006"
    INFO_BLUE = "BDD7EE"
    INFO_BLUE_DARK = "1F4E79"

    # Neutral Colors
    HEADER_BLUE = "4472C4"
    HEADER_DARK = "305496"
    LIGHT_GRAY = "F2F2F2"
    MEDIUM_GRAY = "D9D9D9"
    DARK_GRAY = "595959"

    # Border style
    THIN_BORDER = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )

    MEDIUM_BORDER = Border(
        left=Side(style="medium", color="BFBFBF"),
        right=Side(style="medium", color="BFBFBF"),
        top=Side(style="medium", color="BFBFBF"),
        bottom=Side(style="medium", color="BFBFBF"),
    )

    @classmethod
    def get_title_font(cls) -> Font:
        """Get font for report titles."""
        return Font(bold=True, size=18, color=cls.DARK_GRAY)

    @classmethod
    def get_subtitle_font(cls) -> Font:
        """Get font for section subtitles."""
        return Font(bold=True, size=14, color=cls.DARK_GRAY)

    @classmethod
    def get_header_font(cls) -> Font:
        """Get font for table headers."""
        return Font(bold=True, size=11, color="FFFFFF")

    @classmethod
    def get_header_fill(cls) -> PatternFill:
        """Get fill for table headers."""
        return PatternFill(
            start_color=cls.HEADER_BLUE,
            end_color=cls.HEADER_BLUE,
            fill_type="solid",
        )

    @classmethod
    def get_hpe_green_fill(cls) -> PatternFill:
        """Get HPE green accent fill."""
        return PatternFill(
            start_color=cls.HPE_GREEN,
            end_color=cls.HPE_GREEN,
            fill_type="solid",
        )

    @classmethod
    def get_hpe_green_light_fill(cls) -> PatternFill:
        """Get light HPE green fill."""
        return PatternFill(
            start_color=cls.HPE_GREEN_LIGHT,
            end_color=cls.HPE_GREEN_LIGHT,
            fill_type="solid",
        )

    @classmethod
    def get_success_fill(cls) -> PatternFill:
        """Get success (green) fill."""
        return PatternFill(
            start_color=cls.SUCCESS_GREEN,
            end_color=cls.SUCCESS_GREEN,
            fill_type="solid",
        )

    @classmethod
    def get_warning_fill(cls) -> PatternFill:
        """Get warning (amber) fill."""
        return PatternFill(
            start_color=cls.WARNING_AMBER,
            end_color=cls.WARNING_AMBER,
            fill_type="solid",
        )

    @classmethod
    def get_error_fill(cls) -> PatternFill:
        """Get error (red) fill."""
        return PatternFill(
            start_color=cls.ERROR_RED,
            end_color=cls.ERROR_RED,
            fill_type="solid",
        )

    @classmethod
    def get_info_fill(cls) -> PatternFill:
        """Get info (blue) fill."""
        return PatternFill(
            start_color=cls.INFO_BLUE,
            end_color=cls.INFO_BLUE,
            fill_type="solid",
        )

    @classmethod
    def get_alternate_row_fill(cls) -> PatternFill:
        """Get alternate row fill for zebra striping."""
        return PatternFill(
            start_color=cls.LIGHT_GRAY,
            end_color=cls.LIGHT_GRAY,
            fill_type="solid",
        )

    @classmethod
    def get_center_alignment(cls) -> Alignment:
        """Get center alignment."""
        return Alignment(horizontal="center", vertical="center")

    @classmethod
    def get_left_alignment(cls) -> Alignment:
        """Get left alignment with wrap."""
        return Alignment(horizontal="left", vertical="center", wrap_text=True)

    @classmethod
    def get_number_alignment(cls) -> Alignment:
        """Get right alignment for numbers."""
        return Alignment(horizontal="right", vertical="center")

    @classmethod
    def get_kpi_value_font(cls) -> Font:
        """Get font for KPI values."""
        return Font(bold=True, size=24, color=cls.HPE_GREEN)

    @classmethod
    def get_kpi_label_font(cls) -> Font:
        """Get font for KPI labels."""
        return Font(bold=False, size=10, color=cls.DARK_GRAY)

    @classmethod
    def get_urgency_fill(cls, days_remaining: int | None) -> PatternFill:
        """Get fill color based on urgency (days remaining)."""
        if days_remaining is None:
            return PatternFill()
        if days_remaining <= 7:
            return cls.get_error_fill()
        if days_remaining <= 30:
            return cls.get_warning_fill()
        if days_remaining <= 90:
            return PatternFill(
                start_color="FFF2CC",  # Light yellow
                end_color="FFF2CC",
                fill_type="solid",
            )
        return cls.get_success_fill()

    @classmethod
    def get_status_fill(cls, status: str | None) -> PatternFill:
        """Get fill color based on status string."""
        if not status:
            return PatternFill()

        status_lower = status.lower()
        if status_lower in ("success", "completed", "active", "online", "started", "assigned", "good"):
            return cls.get_success_fill()
        if status_lower in ("warning", "pending", "expiring", "partial", "fair"):
            return cls.get_warning_fill()
        if status_lower in ("error", "failed", "offline", "ended", "cancelled", "poor"):
            return cls.get_error_fill()
        if status_lower in ("info", "running", "syncing"):
            return cls.get_info_fill()
        return PatternFill()

    @classmethod
    def create_named_styles(cls) -> list[NamedStyle]:
        """Create named styles for the workbook."""
        styles = []

        # Header style
        header_style = NamedStyle(name="report_header")
        header_style.font = cls.get_header_font()
        header_style.fill = cls.get_header_fill()
        header_style.alignment = cls.get_center_alignment()
        header_style.border = cls.THIN_BORDER
        styles.append(header_style)

        # Data cell style
        data_style = NamedStyle(name="report_data")
        data_style.font = Font(size=10)
        data_style.alignment = cls.get_left_alignment()
        data_style.border = cls.THIN_BORDER
        styles.append(data_style)

        # Number style
        number_style = NamedStyle(name="report_number")
        number_style.font = Font(size=10)
        number_style.alignment = cls.get_number_alignment()
        number_style.border = cls.THIN_BORDER
        number_style.number_format = "#,##0"
        styles.append(number_style)

        # Percentage style
        percent_style = NamedStyle(name="report_percent")
        percent_style.font = Font(size=10)
        percent_style.alignment = cls.get_number_alignment()
        percent_style.border = cls.THIN_BORDER
        percent_style.number_format = "0.0%"
        styles.append(percent_style)

        # Date style
        date_style = NamedStyle(name="report_date")
        date_style.font = Font(size=10)
        date_style.alignment = cls.get_center_alignment()
        date_style.border = cls.THIN_BORDER
        date_style.number_format = "YYYY-MM-DD"
        styles.append(date_style)

        return styles
